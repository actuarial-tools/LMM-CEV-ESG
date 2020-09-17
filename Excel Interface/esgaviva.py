# -*- coding: utf-8 -*-
"""
Created on Tue May 19 16:46:32 2020

@author: Caleb
"""

# Extra Packages
from openpyxl import load_workbook
import re
import scipy.optimize as opt
from lmfit import minimize, Parameters, Parameter, report_fit
from itertools import compress
import seaborn as sns
from sklearn.decomposition import PCA
from zipfile import ZipFile
import xlwings as xlw
from datetime import datetime
from nelson_siegel_svensson import NelsonSiegelSvenssonCurve
from nelson_siegel_svensson.calibrate import calibrate_nss_ols
import dask.dataframe as dk


# General management
import numpy as np
from math import *
import pandas as pd
from scipy import stats, optimize
import itertools
import matplotlib.pyplot as plt
import sobol_seq

from time import strftime, time
import copy

# Pretty Table
from prettytable import PrettyTable

# Manage operating systems
import os

# General use packages
from random import seed # Sets the seed

import sys

loc = '//cerata/DGCU/Risques/16. Actuarial Function/05. Economic Scenario Générator/00 - DD LMM CEV/02 - Travaux/Excel Interface'
os.chdir(loc)

'''DEFINE ALL LOCATIONS USED FOR INPUTS AND OUTPUTS'''
# 1. Volatility Surface
volSurfaceLocation = "default_data\\Volatility Surface.xlsx"

# 2. EIOPA Curve
curveLocation = "default_data\\EIOPA Curve.xlsx"

# 3. Weights
weightsLocation = "default_data\\Weights.xlsx"

# 4. Betas
betasLocation = "default_data\\Betas YE19.xlsx"

# 5. Results
resultsLocation = "default_data\\Results.xlsx"

# 6. Gaussians
gaussiansLocation = "default_data\\Gaussians.xlsx"

# 7 Rates Distributions
ratesElasticityLocation = "default_data\\Rates Elasticity = 0.8.xlsx"

# 8 ZC Distributions
zcElasticityLocation = "default_data\\ZeroCoupons Elasticity = 0.8.xlsx"

# 9 Exported Gaussians1
exportedGaussian1Location = "default_data\\ExportedGaussians1.xlsx"

# 10 Exported Gaussians2
exportedGaussian2Location = "default_data\\ExportedGaussians2.xlsx"


''' CALIBRATION AND PROJECTION INPUTS
==================================================================================================='''

'''EIOPA CURVE
------------------------------------'''
EIOPACurve = np.array(pd.read_excel(curveLocation,
              header = None,
             index_col=0))

# Convert Rates to ZC Prices {(1/(1 + rate))^maturity}
EIOPAZCCurve = [pow(1/(1+EIOPACurve[i-1]), i) for i in range(1, len(EIOPACurve)+1)]

# Obtain the vol surface {(ZC(Tk-1)/ZC(Tk))-1}
EIOPAForwardCurve = [(EIOPAZCCurve[i-1]/EIOPAZCCurve[i])-1 for i in range(1, len(EIOPAZCCurve))]

# Define the module ZC and Forward Curves
'Note that this curve will be used for the entirety of the model, to change it, please change the input EIOPA Curve.xlsx in the Inputs Folder'
zeroCouponCurve = np.array(EIOPAZCCurve)
forwardCurve = np.array(EIOPAForwardCurve)


'''Barrie & Hibbert betas'''
betas = pd.read_excel(betasLocation).transpose().to_numpy()


'Simple Smith Wilson Implementation'
'----------------------------------------'
class SmithWilson:
    maturities = np.arange(1, 41) #These are the default input maturities
    
    def __init__(self, ZCPrices, maturities = maturities, alpha = 0.134,  UFR = 0.039):
        self.UFR = np.log(1+UFR)
        self.alpha = alpha
        self.maturities = maturities
        self.ZCPrices = ZCPrices
    
    'Define the Wilson Function'
    def wilsonFunction(self, t, uj):
        mu = np.exp(- self.UFR*(t+uj))
        maxT = self.alpha * np.maximum(t, uj)
        minT = self.alpha * np.minimum(t, uj)
        
        # We define the Wilson kernel function
        wj = mu * (minT - 0.5 * np.exp(-maxT)*(np.exp(minT) - np.exp(-minT)))
        
        return (wj)
    
    def wilsonVectorFunction(self, inputMaturities):
        wilsonMatrix = np.zeros((len(inputMaturities), len(self.maturities))) 
        for t in inputMaturities:
            for j in self.maturities:
                wilsonMatrix[t-1, j-1] = self.wilsonFunction(t, j)
            
        return(wilsonMatrix)
    
    'Obtain the parameter vector zeta'
    def calibrate(self):
        #Create the matrix W of Kernels
        self.WMatrix = self.wilsonVectorFunction(self.maturities)
                
        #Invert the matrix
        #Recall that our parameters are W^-1 * (ZCPrices - muVector)
        WMatrixInv = np.linalg.inv(self.WMatrix) # Invert the kernel
        muVector = np.exp(-self.UFR * self.maturities)
        
        SWParams = WMatrixInv.dot(self.ZCPrices  - muVector) # zeta = W(inverse) * (ZCprice - mu)
        
        return (SWParams)
    
    'Fit Curve'
    def curve(self):
        params = self.calibrate() #Obtain parameters
        parametrizedWilson = self.wilsonVectorFunction(np.arange(1, 151)).dot(params) # Fit full EIOPA Curve
        result = np.exp(-self.UFR * np.arange(1, 151)) + parametrizedWilson
                    
        return(result)
       



'''
**********************************************************************************************************
*                                       II) CALIBRATION                                                  *
**********************************************************************************************************
'''
'''PRICERS CALIBRATION
========================='''
def forwardSwapRateCalib(expiry, tenor):
    '''Quick Calculation of the forward swap rate based on the module Zero Coupon Curve'''
    num = zeroCouponCurve[expiry-1] - zeroCouponCurve[expiry+tenor-1]
    
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
    return(num/PVBP)

#Black Pricer - Payer
######################################
def blackPayerShifted(N, S0, K, expiry, tenor, sigma, delta):
    '''Calculation of the swaption price based on the Shifted Black Model
    Inputs
    =======
    N - Notional
    S0 - Forward Swap Rate
    K - Strike
    expiry, tenor, sigma are self explanatory
    delta = shift on the Black Model'''     
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
    # Define Black Parameters
    T = expiry
    d1 = (log((S0+delta)/(K + delta)) + 0.5*(pow(sigma,2)*T))/(sigma * sqrt(T))
    d2 = d1 - sigma*sqrt(T)
    
    price = N*PVBP * ((S0+delta)*stats.norm.cdf(d1) - (K+delta)*stats.norm.cdf(d2))
    return(price)


# Normal Pricer - Payer
######################################
def normalPayer(N, S0, K, expiry, tenor, sigma):
    '''Calculation of the payer swaption price based on the Bachelier Model
    Inputs
    =======
    N - Notional
    S0 - Forward Swap Rate
    K - Strike
    expiry, tenor, sigma are self explanatory'''  
    
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
     # Define Black Parameters
    T = expiry
    d1 = (S0 - K)/(sigma * sqrt(T))
    
    #Calculate Price
    price = N * PVBP * ((S0 - K)*stats.norm.cdf(d1) + (sigma * sqrt(T)* exp(-0.5 * pow(d1, 2)))/sqrt(2 * pi))
    return(price)

# Normal Pricer - Receiver
######################################
def normalReceiver(N, S0, K, expiry, tenor, sigma):
    '''Calculation of the receiver swaption price based on the Bachelier Model
    Inputs
    =======
    N - Notional
    S0 - Forward Swap Rate
    K - Strike
    expiry, tenor, sigma are self explanatory'''
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
     # Define Black Parameters
    T = expiry
    d1 = (S0 - K)/(sigma * sqrt(T))
    
    #Calculate Price
    price = N * PVBP * ((K - S0)*stats.norm.cdf(-d1) + (sigma * sqrt(T)* exp(-0.5 * pow(d1, 2)))/sqrt(2 * pi))
    return(price)

# Black Shifted - Payer
######################################
def blackShifted(N, S0, K, expiry, tenor, sigma, delta):
    '''Calculation of the ATM payer swaption price based on the Shifted Black Model
    Inputs
    =======
    N - Notional
    S0 - Forward Swap Rate
    K - Strike
    expiry, tenor, sigma are self explanatory
    delta = Shift on the Black Model'''
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
    # Define Black Parameters
    T = expiry
    d1 = (log((S0+delta)/(K + delta)) + 0.5*(pow(sigma,2)*T))/(sigma * sqrt(T))
    d2 = d1 - sigma*sqrt(T)
    
    price = N * PVBP * ((S0+delta)*stats.norm.cdf(d1) - (K+delta)*stats.norm.cdf(d2))
    return(price)

# Black Shifted - Payer
######################################
def blackShiftedCalib(expiry, tenor, sigma, delta):
    '''Calculation of the ATM payer swaption price based on the Shifted Black Model
    Inputs
    =======
    expiry, tenor, sigma are self explanatory
    delta = shift on the Black Model'''
    S0 = forwardSwapRateCalib(expiry, tenor)
    K = forwardSwapRateCalib(expiry, tenor)
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
    # Define Black Parameters
    T = expiry
    d1 = (log((S0+delta)/(K + delta)) + 0.5*(pow(sigma,2)*T))/(sigma * sqrt(T))
    d2 = d1 - sigma*sqrt(T)
    
    price = PVBP * ((S0+delta)*stats.norm.cdf(d1) - (K+delta)*stats.norm.cdf(d2))
    return(price)


'''IMPLICIT (BACHELIER) VOLATILITY FUNCTIONS
=============================================='''
# Newton Algorithm
def imp_vol_payer_Newton(N, S0, K, expiry, tenor,Price):
    '''Use the classical Newton Algorithm to find the implied Black vol of a payer swaption
    Inputs
    =======
    expiry, tenor, sigma are self explanatory
    delta = shift on the Black Model
    Price = Swaption Price'''
    
    func = lambda sigma: np.power(normalPayer(N, S0, K, expiry, tenor, sigma) - Price, 2.0)
    vol  = optimize.newton(func, 0.1, tol = 1e-5, maxiter=100000)
    return(vol)

# Nelder Mead Algorithm
def imp_vol_payer_NMead(N, S0, K, expiry, tenor,Price):
    '''Use the classical Nelder Mead Algorithm to find the implied Black vol of a payer swaption
    Inputs
    =======
    expiry, tenor, sigma are self explanatory
    delta = shift on the Black Model
    Price = Swaption Price'''
    
    func = lambda sigma: np.power(normalPayer(N, S0, K, expiry, tenor, sigma) - Price, 2.0)
    vol  = optimize.minimize(func, 0.1, method='nelder-mead',
                             options={'xatol': 1e-6, 'disp': False, 'maxiter':100000})
    return(vol.x)

# Toms748 Algorithm
def imp_vol_payer_toms748(N, S0, K, expiry, tenor,Price):
    '''Use the classical Toms748 Algorithm to find the implied Black vol of a payer swaption
    Inputs
    =======
    expiry, tenor, sigma are self explanatory
    delta = shift on the Black Model
    Price = Swaption Price'''
    func = lambda sigma: np.power(normalPayer(N, S0, K, expiry, tenor, sigma) - Price, 1.0)
    vol  = optimize.toms748(func,a = -1, b = 1, xtol = 1e-5)
    return(vol)

def volNormalATMFunction(expiry, tenor, price):
    '''Use the classical Nelder Mead Algorithm to find the implied Bachelier vol of a payer swaption
    Inputs
    =======
    expiry, tenor are self explanatory
    Price = Swaption Price'''
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
    volNormal = (price /PVBP) * np.sqrt(2 * pi/expiry)
    return (volNormal)


'''VECTORIZING ALL FUNCTIONS
======================================'''
Toms_Vect = np.vectorize(imp_vol_payer_toms748)
Newton_Vect = np.vectorize(imp_vol_payer_Newton)
NMead_Vect = np.vectorize(imp_vol_payer_NMead)
blackVect = np.vectorize(blackShifted)
volNormalVect = np.vectorize(volNormalATMFunction)
blackPayerShiftedVect = np.vectorize(blackPayerShifted)
forwardSwapRateVect = np.vectorize(forwardSwapRateCalib)
normalPayerVect = np.vectorize(normalPayer)
blackVectCalib = np.vectorize(blackShiftedCalib)
volNormalATMFunctionVect = np.vectorize(volNormalATMFunction)


'''VOLATILITY SURFACE
=====================================
- Some of the functions above are required in the construction of the volatility data
- The forwardSwapRateCalib function is used to calculate strikes
- The normalPayerVect function has been used to calculate the normal Prices based on our surface

'''
'''VOLATILITY SURFACE
------------------------------------'''
volatilitiesLMMPlus = pd.read_excel(volSurfaceLocation, 
                    index_col = 0).unstack().reset_index(name='value')

# Import square/rectangle dataset and transform it to a 3 columns with Expiry, Tenor and the Weight
weightsLMMPlus = pd.read_excel(weightsLocation, 
                               index_col = 0).unstack().reset_index(name='value')

# Name the columns
volatilitiesLMMPlus.columns = ['Tenor', 'Expiry', 'Value']
weightsLMMPlus.columns = ['Tenor', 'Expiry', 'Value']

# Unintuitive naming but used for Martingale and Monte Carlo Tests
expiriesBachelier = weightsLMMPlus['Expiry'].to_numpy() 
tenorsBachelier = weightsLMMPlus['Tenor'].to_numpy()

# Calculate the forward Swap Rates for each Expiry X Tenor couple
strikesLMMPlus = forwardSwapRateVect(weightsLMMPlus['Expiry'],weightsLMMPlus['Tenor'] )

# Calculate the Bachelier Price
normalPricesLMMPlus = normalPayerVect(1, strikesLMMPlus, strikesLMMPlus,
                                      weightsLMMPlus['Expiry'],
                                      weightsLMMPlus['Tenor'],
                                      volatilitiesLMMPlus['Value'])  



'''SIGMA ALPHA BETA FUNCTION
=============================
1) PCA parametrization'''
def sigmaAlphaBetaPCA(expiry, tenor, fZero, gamma, a, b, c, d, eta, delta):
    '''Calculates the sigmaAlphaBeta input in the Chi Square pricer
        (Please refer to Andreasen & Andersen (2002)
        
        Inputs
        =========
        expiry, tenor are self explanatory
        eta = Elasticity of the CEV Model
        delta = shift on the model
        
        Inputs from the PCA volatility parametrization:
        (fZero, gamma, a, b, c d )
        
        Please reference the PCA volatility description in the documentation of the DD LFM CEV.'''
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:tenor+expiry])

    # Calculate the forward swap rate
    swapForward = forwardSwapRateCalib(expiry, tenor)

    # Calculate the ZC bond value at T_alpha and T_beta
    zcAlpha = zeroCouponCurve[expiry-1]
    zcBeta = zeroCouponCurve[expiry+tenor-1]

    # Calculate weights
    '''To be confirmed
    ==================='''
    weights = [(swapForward*np.power(forwardCurve[i-1]+ delta, eta)/
                     (np.power(swapForward + delta, eta)*(1 + forwardCurve[i-1])))* 
                    ((zcBeta/(zcAlpha-zcBeta)) + np.sum(zeroCouponCurve[i:expiry+tenor])/PVBP)
                    for i in np.arange(expiry, expiry+tenor)]

    '''VOLATILITY FUNCTIONS
    ========================'''
    # f and g vectors 
    fValues = fZero + (1 - fZero)*np.exp(-gamma * np.arange(tenor + expiry))
    gValues = ((a + b * np.arange(1, tenor + expiry + 1))* 
                np.exp(-c * np.arange(1, tenor + expiry + 1))) + d
    
    # Product of g and betas (Since they have the same index)
    gBeta1 = gValues * betas[0, :tenor+expiry]
    gBeta2 = gValues * betas[1, :tenor+expiry]

    '''INTEGRAL
    ============'''
    sigmaIntegral1 = np.sum([np.square(np.sum(fValues[i]*gBeta1[expiry-i-1:expiry+tenor-i-1]*weights))
                                for i in range(expiry)])
    sigmaIntegral2 = np.sum([np.square(np.sum(fValues[i]*gBeta2[expiry-i-1:expiry+tenor-i-1]*weights))
                                for i in range(expiry)]) 
    
    return(sigmaIntegral1 + sigmaIntegral2)


'''SIGMA ALPHA BETA FUNCTION: Hagan
===================================='''
def sigmaAlphaBetaHagan(expiry, tenor, fZero, gamma, a, b, c, d, eta, delta):
    '''Calculates the sigmaAlphaBeta input in the Hagan Pricer
        (Please refer to Andreasen & Andersen (2002) and Hagan Woodward 
        documentation on CEV approximation(1998))
        
        Inputs
        =========
        expiry, tenor are self explanatory
        eta = Elasticity of the CEV Model
        delta = shift on the model
        
        Inputs from the PCA volatility parametrization:
        (fZero, gamma, a, b, c d )
        
        Please reference the PCA volatility description in the documentation of the DD LFM CEV.'''
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:tenor+expiry])
    
    # Calculate the forward swap rate
    swapForward = forwardSwapRateCalib(expiry, tenor)

    # Calculate the ZC bond value at T_alpha and T_beta
    zcAlpha = zeroCouponCurve[expiry-1]
    zcBeta = zeroCouponCurve[expiry+tenor-1]
    
    # Calculate weights
    weights = [(swapForward*np.power(forwardCurve[i-1]+ delta, eta)/
                 (np.power(swapForward + delta, eta)*(1 + forwardCurve[i-1])))* 
                ((zcBeta/(zcAlpha-zcBeta)) + np.sum(zeroCouponCurve[i:expiry+tenor])/PVBP)
                for i in np.arange(expiry, expiry+tenor)]
    
    '''VOLATILITY FUNCTIONS
    ========================'''
    # f and g vectors 
    fValues = fZero + (1 - fZero)*np.exp(-gamma * np.arange(tenor + expiry)) # Index begins at 0
    gValues = ((a + b * np.arange(1, tenor + expiry + 1))* 
                np.exp(-c * np.arange(1, tenor + expiry + 1))) + d # Index begins at 1
    
    # Product of g and betas (Since they have the same index)
    gBeta1 = gValues * betas[0, :tenor+expiry]
    gBeta2 = gValues * betas[1, :tenor+expiry]

    '''INTEGRAL
    ============'''
    sigmaIntegral1 = np.sum([np.square(np.sum(fValues[i]*gBeta1[expiry-i-1:expiry+tenor-i-1] * weights))
                                for i in range(expiry)])
    sigmaIntegral2 = np.sum([np.square(np.sum(fValues[i]*gBeta2[expiry-i-1:expiry+tenor-i-1] * weights))
                                for i in range(expiry)]) 
 
    return(np.sqrt((sigmaIntegral1 + sigmaIntegral2)/expiry))


''' Hagan Approximation 
=============================='''
def sigma_CEV_Hagan(expiry, tenor, fZero, gamma, a, b, c, d, eta, delta):
    '''Calculates the equivalent Black volatility using the Hagan approximation
    (Please refer to Hagan Woodward documentation on CEV approximation(1998))
    
    Inputs
    =========
    expiry, tenor are self explanatory
    eta = Elasticity of the CEV Model
    delta = shift on the model
    
    Inputs from the PCA volatility parametrization:
    (fZero, gamma, a, b, c d )
    
    Please reference the PCA volatility description in the documentation of the DD LFM CEV.'''
    sigma =  sigmaAlphaBetaHagan(expiry, tenor, fZero, gamma, a, b, c, d, eta, delta)
    Sav = forwardSwapRateCalib(expiry, tenor) + delta
    
    term1 = sigma/np.power(Sav, 1 - eta)
    term4 = (np.power((1 - eta)*sigma,2)*expiry)/(24 * np.power(Sav,2 - 2*eta))
   
    sigma_black = term1 * (1 + term4)
    
    return(sigma_black)


'''CHI SQUARE PAYER WITH PYTHON 
================================='''
def calibrationChiSquarePayerPython(N, S0, K, expiry, tenor, fZero, gamma, a, b, c, d, eta, delta):
    '''Calculates the swaption Payer prices using the Andreasen & Andersen swaption pricing formula
    Uses Python to generate chi square CDF values 
    
    (Please refer to Andreasen & Andersen CEV swaption pricing formula (2002))
    
    Inputs
    =========
    expiry, tenor are self explanatory
    eta = Elasticity of the CEV Model
    delta = shift on the model
    
    Inputs from the PCA volatility parametrization:
    (fZero, gamma, a, b, c d )
    
    Please reference the PCA volatility description in the documentation of the DD LFM CEV.
    
    NOTE
    =====
    * Function produces negative values for swaptions with very tenor = 1 and 2'''
    S0 = forwardSwapRateCalib(expiry, tenor)
    K = forwardSwapRateCalib(expiry, tenor)
    
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
    # Parameters for input into Chi Square CDF
    sigmaSquaredT =  sigmaAlphaBetaPCA(expiry, tenor, fZero, gamma, a, b, c, d, eta, delta)
    d = (pow(K+delta,2 - 2*eta))/(pow(1 - eta,2)* sigmaSquaredT)
    b = 1/(1 - eta)
    f = (pow(S0+delta, 2 - 2*eta))/(pow(1 - eta, 2)* sigmaSquaredT)

    # Calculate Price
    price = PVBP * ((S0+delta)*(1 - stats.ncx2.cdf(d, b+2, f)) - (K+delta)*(stats.ncx2.cdf(f, b, d)))
    return(price)



''' Black Pricer - Payer
=============================='''
def calibrationBlackHaganPayer(N, S0, K, expiry, tenor, fZero, gamma, a, b, c, d, eta, delta):    
    '''Calculates the swaption Payer prices using the Hagan approximation 
    (Please refer to Hagan Woodward documentation on CEV approximation(1998))
    
    Inputs
    =========
    expiry, tenor are self explanatory
    eta = Elasticity of the CEV Model
    delta = shift on the model
    
    Inputs from the PCA volatility parametrization:
    (fZero, gamma, a, b, c d )
    
    Please reference the PCA volatility description in the documentation of the DD LFM CEV.'''     
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
    sigma = sigma_CEV_Hagan(expiry, tenor, fZero, gamma, a, b, c, d, eta, delta)
    
    # Define Black Parameters
    T = expiry
    d1 = (log((S0+delta)/(K+delta)) + 0.5*(pow(sigma,2)*T))/(sigma * sqrt(T))
    d2 = d1 - sigma*sqrt(T)
    
    price = PVBP * ((S0+delta)*stats.norm.cdf(d1) - (K+delta)*stats.norm.cdf(d2))
    return(price)



'''CALCULATE THE MARKET PRICES
==================================================================='''
# Vectorize Function
calibrationBlackVect = np.vectorize(calibrationBlackHaganPayer)
sigma_CEV_Hagan_Vect = np.vectorize(sigma_CEV_Hagan)
sigmaAlphaBetaHaganVect = np.vectorize(sigmaAlphaBetaHagan)
sigmaAlphaBetaPCAVect = np.vectorize(sigmaAlphaBetaPCA)
calibrationChiSquarePayerPythonVect = np.vectorize(calibrationChiSquarePayerPython)


'''OPTIMIZATION FUNCTIONS (LEVENBERG MARQUARDT IMPLEMENTATION)
=================================================================='''

# Calibration Function (Python Version)
#############################################
def chiSquareCalibratorPython(initialValues = [0.1,0.1,0.1,0.1,0.1,0.1], 
                              lowerBounds = [0,0,0,0,0,0], 
                              upperBounds = [1,1,1,1,1,1], eta = 0.8, delta = 0.1, iter_cb = None):
    '''Implementation of the Levenberg Marquardt algorithm in Python to find the optimal value 
        based on a given volatility surface.
        
        Function to be minimized:
            Error = (MarketPrices - ModelPrices)/MarketPrices
        
        INPUTS
        ===========
        1) Volatility Surface
            - Obtained from the Excel File in the LMM Volatility Surface.xlsx in the Inputs File
        
        2) EIOPA Curve
            - Obtained from the Excel File in the EIOPA Curve.xlsx in the Inputs File
            
        3) initialValues
            - Initialization values for the algorithms in this order:
                [fZero, gamma, a, b, c, d]
                
            - Default value: [0.1,0.1,0.1,0.1,0.1,0.1]
            
        4) lowerBounds
            -Fix lower limit for the values
            - Default value: [0.0,0.0,0.0,0.0,0.0,0.0]
            
        5) upperBounds
            -Fix upper limit for the values
            - Default value: [1.0,1.0,1.0,1.0,1.0,1.0]
            
        6) eta is the elasticity metaparameter and delta the shift.
    
    NOTE
    =======
    1) Tolerances have all been set to 1e-19 focusing more on precision than runtime
    '''
        
    '''DEFINE THE PARAMETERS 
    =========================='''
    params = Parameters()
    params.add('fZero', value = initialValues[0], min = lowerBounds[0], max = upperBounds[0])
    params.add('gamma',value = initialValues[1], min = lowerBounds[1], max = upperBounds[1])
    params.add('a',value = initialValues[2], min = lowerBounds[2], max = upperBounds[2])
    params.add('b',value = initialValues[3], min = lowerBounds[3], max = upperBounds[3])
    params.add('c', value = initialValues[4], min = lowerBounds[4], max = upperBounds[4])
    params.add('d', value = initialValues[5], min = lowerBounds[5], max = upperBounds[5])
    
    # Objective Function
    objectiveFunctionLMMPlus = lambda test: weightsLMMPlus['Value'].to_numpy() *  ( normalPricesLMMPlus - 
        calibrationChiSquarePayerPythonVect(1.0, strikesLMMPlus, 
                        strikesLMMPlus, 
                        weightsLMMPlus['Expiry'], 
                        weightsLMMPlus['Tenor'], 
                        test['fZero'].value, 
                        test['gamma'].value, 
                        test['a'].value, 
                        test['b'].value,
                        test['c'].value,
                        test['d'].value,
               eta, delta))/normalPricesLMMPlus       
    
   
    '''OPTIMIZE PARAMETERS PYTHON
    =============================='''
    try:
        result = minimize(objectiveFunctionLMMPlus, 
                          params, 
                          method = 'leastsq',
                           iter_cb = iter_cb, 
                          ftol = 1e-12,
                          xtol = 1e-12,
                          gtol = 1e-12)
    except:
        print('')
        print('CALIBRATION ERROR ENCOUNTERED')
        print('----------------------------------')
        print('Possible errors:')
        print('    i)   Incorrect format of inputs (vol surface or weights.')
        print('    ii)  Boundaries & Inputs should be in NUMERIC format.')
        print('    iii) Saturation of algorithm/Impossible to find solution.')
           
    finalResults = [result.params['fZero'].value, 
            result.params['gamma'].value, 
            result.params['a'].value, 
            result.params['b'].value, 
            result.params['c'].value,
            result.params['d'].value, eta, delta]
    
    return(finalResults)

# Calibration Function (Python Version)
#############################################
def haganCalibratorPython(initialValues = [0.1,0.1,0.1,0.1,0.1,0.1], 
                              lowerBounds = [0,0,0,0,0,0], 
                              upperBounds = [1,1,1,1,1,1], eta = 0.8, delta = 0.1, iter_cb = None):
    '''Implementation of the Levenberg Marquardt algorithm in Python to find the optimal value 
        based on a given volatility surface.
        
        Model used = Hagan Approximation
        
        Function to be minimized:
            Error = (MarketPrices - ModelPrices)/MarketPrices
        
        INPUTS
        ===========
        1) Volatility Surface
            - Obtained from the Excel File in the LMM Volatility Surface.xlsx in the Inputs File
        
        2) EIOPA Curve
            - Obtained from the Excel File in the EIOPA Curve.xlsx in the Inputs File
            
        3) initialValues
            - Initialization values for the algorithms in this order:
                [fZero, gamma, a, b, c, d]
                
            - Default value: [0.1,0.1,0.1,0.1,0.1,0.1]
            
        4) lowerBounds
            -Fix lower limit for the values
            - Default value: [0.0,0.0,0.0,0.0,0.0,0.0]
            
        5) upperBounds
            -Fix upper limit for the values
            - Default value: [1.0,1.0,1.0,1.0,1.0,1.0]
            
        6) eta is the elasticity metaparameter and delta the shift.
    
    NOTE
    =======
    1) Tolerances have all been set to 1e-19 focusing more on precision than runtime
    '''
        
    '''DEFINE THE PARAMETERS 
    =========================='''
    params = Parameters()
    params.add('fZero', value = initialValues[0], min = lowerBounds[0], max = upperBounds[0])
    params.add('gamma',value = initialValues[1], min = lowerBounds[1], max = upperBounds[1])
    params.add('a',value = initialValues[2], min = lowerBounds[2], max = upperBounds[2])
    params.add('b',value = initialValues[3], min = lowerBounds[3], max = upperBounds[3])
    params.add('c', value = initialValues[4], min = lowerBounds[4], max = upperBounds[4])
    params.add('d', value = initialValues[5], min = lowerBounds[5], max = upperBounds[5])
    
    # Objective Function
    objectiveFunctionLMMPlus = lambda test: weightsLMMPlus['Value'].to_numpy() *  ( normalPricesLMMPlus - 
        calibrationBlackVect(1.0, strikesLMMPlus, 
                        strikesLMMPlus, 
                        weightsLMMPlus['Expiry'], 
                        weightsLMMPlus['Tenor'], 
                        test['fZero'].value, 
                        test['gamma'].value, 
                        test['a'].value, 
                        test['b'].value,
                        test['c'].value,
                        test['d'].value,
               eta, delta))/normalPricesLMMPlus       
    
   
    '''OPTIMIZE PARAMETERS PYTHON
    =============================='''
    try:
        result = minimize(objectiveFunctionLMMPlus, 
                          params, 
                          method = 'leastsq',
                          iter_cb = iter_cb, 
                          ftol = 1e-12,
                          xtol = 1e-12,
                          gtol = 1e-12)
    except:
        print('')
        print('CALIBRATION ERROR ENCOUNTERED')
        print('----------------------------------')
        print('Possible errors:')
        print('    i)   Incorrect format of inputs (vol surface or weights.')
        print('    ii)  Boundaries & Inputs should be in NUMERIC format.')
        print('    iii) Saturation of algorithm/Impossible to find solution.')
           
    finalResults = [result.params['fZero'].value, 
            result.params['gamma'].value, 
            result.params['a'].value, 
            result.params['b'].value, 
            result.params['c'].value,
            result.params['d'].value, eta, delta]
    
    return(finalResults)

def fullCalibratorPython(initialValues = [0.1,0.1,0.1,0.1,0.1,0.1], 
                              lowerBounds = [0,0,0,0,0,0], 
                              upperBounds = [1,1,1,1,1,1], eta = 0.8, delta = 0.1, 
                              iter_chi = None,
                              iter_Hagan = None):
    
    resultsChi2 = chiSquareCalibratorPython(initialValues, 
                                            lowerBounds, 
                                            upperBounds,
                                            eta, delta, iter_chi)

    resultsHagan = haganCalibratorPython(initialValues, 
                                            lowerBounds, 
                                            upperBounds,
                                            eta, delta, 
                                            iter_Hagan)
    
    return([resultsChi2, resultsHagan])


'''
**********************************************************************************************************
*                                       III) PROJECTION                                                  *
**********************************************************************************************************
'''

'''
1) BH INPUT BROWNIENS
===================================================================================================
'''
def BHSingleSimulator(forwardCurve, fZero, gamma, a, b, c, d, eta, betas, delta, dZ1, dZ2, 
                      maxProjectionYear = 50, maxMaturity = 40):
    '''Simulates a single entire forward curve scenario using Moodys gaussians.
    
    Inputs
    ========
    forwardCurve = Curve intended to be projected. Select the curve length depending on the maximum maturity intended for projection.
    
    dZ1, dZ2 = Gaussians obtained from the Moodys tool
    
    fZero, gamma, a, b, c, d = parameters for the volatility parametrization
    
    NOTE
    =======
    1) Simulation set by default to 50 years. This is what is used in the internal model
    2) No need to define the simulations
    3) This simulation is not just constrained to Moodys Gaussians, any Gaussians in the correct will work
        -For formatting, please check the BH Browniens.xlsx file in the Inputs folder
    '''
    timeSteps = len(forwardCurve[:maxMaturity+maxProjectionYear]) 
    dt = 1
      
    # Set the forward Curve
    curve = np.array(forwardCurve[:maxMaturity+maxProjectionYear])
    qCurve = np.power(curve+delta, 1 - eta)/(1 - eta)
    
    # f and g vectors 
    fValues = fZero + (1 - fZero)*np.exp(-gamma * np.arange(timeSteps))
    gValues = ((a + b * np.arange(1, timeSteps+1)) * np.exp(-c * np.arange(1, timeSteps+1))) + d
    
    # Product of g and betas (Since they have the same index)
    gBeta1 = gValues * betas[0, :timeSteps]
    gBeta2 = gValues * betas[1, :timeSteps]
    
    terminalCurve = []
    deflateurs = []
    gaussians = np.array([dZ1, dZ2]).transpose()

    
    # Loop to construct curves
    for t in range(1, maxProjectionYear+1):  
        # Calculate Drift
        mu1 = np.array(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*
                             np.cumsum(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]* np.power(curve + delta, eta)/
                                   (1 + curve)))       
        mu2 = np.array(fValues[t-1]*gBeta2[:len(gBeta2)-t+1] *
                             np.cumsum(fValues[t-1]*gBeta2[:len(gBeta2)-t+1]* np.power(curve + delta, eta)/
                                   (1 + curve)))
        mu = mu1 + mu2

        fCurveSum = np.power(curve + delta, eta-1)
        
        # Recalculate the qCurve as it should be
        # Construct Curve     
        qCurve = (qCurve
                    + dt * (mu 
                    - 0.5 *eta*  fCurveSum * (np.power(fValues[t-1]*gBeta1[:len(gBeta1)-t+1], 2) + 
                                 np.power(fValues[t-1]*gBeta2[:len(gBeta2)-t+1], 2)))
                    + fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*np.sqrt(dt)*gaussians[t-1, 0] 
                    + fValues[t-1]*gBeta2[:len(gBeta2)-t+1]*np.sqrt(dt)*gaussians[t-1, 1] )
        
        # Since the timestep is 1 we do this for each simulation
        terminalCurve.append(np.power((1-eta)*qCurve, 1/(1-eta)) - delta)
        deflateurs.append(1/(1 + (np.power((1-eta)*qCurve[0], 1/(1-eta))) - delta))
        
        qCurve = qCurve[1:] # Delete 1 year from curve
        curve = np.power((1-eta)*qCurve, 1/(1-eta)) - delta
                    
    return([terminalCurve,
            deflateurs])


'''
2) DEFINING THE FINAL PROJECTOR FUNCTION
===================================================================================================
'''
def BHfullSimulator(forwardCurve, fZero, gamma, a, b, c, d, eta, betas, delta, 
                    maxProjectionYear = 50, maxMaturity = 40, 
                    inputLocation = gaussiansLocation):
    '''Loop to simulate forward curve scenarios using Moodys gaussians.
        Based on the BHSingleSimulator function
    
    Inputs
    ========
    forwardCurve = Curve intended to be projected. Select the curve length depending on the maximum maturity intended for projection.
    
    fZero, gamma, a, b, c, d = parameters for the volatility parametrization
    
    zeroCouponMGTest = Binary value allowing for Martingale Tests to be conducted on the generated Scenarios. 
    
    SwaptionMonteCarloTest = Binary value allowing for Swaption Monte Carlo Tests
      
    Outputs
    ========
    1) simulatedCurves & deflateurs = All interest rate scenarios
    2) MGTests = all saved in the Results.xlsx file in the Outputs Folder
    3) Zero Coupon scenarios are all saved in the ZCScenarios.csv file in the Outputs folder
    
    NOTE
    =======
    1) Simulation set by default to 50 years. This is what is used in the internal model
    2) No need to define the simulations
    3) This simulation is not just constrained to Moodys Gaussians, any Gaussians in the correct will work
        -For formatting, please check the BH Browniens.xlsx file in the Inputs folder
    '''
    try:
        start = time()
        # Import data from BH Brownians
        browniansBH = pd.read_excel(inputLocation, 
                                    names = ['Trial', 'Timestep', 'Gaussian1', 'Gaussian2'], 
                                    index_col = [0,1])
        maxCol = np.amax(np.array(browniansBH.index.get_level_values(1)))
        simulations = int(len(browniansBH)/maxCol)
        browniensBH1 = browniansBH['Gaussian1'].to_numpy().reshape((simulations, maxCol))
        browniensBH2 = browniansBH['Gaussian2'].to_numpy().reshape((simulations, maxCol))
        
        # Where results will be saved
        simulatedCurves = []
        deflateurs = []
        
        for i in np.arange(simulations, step = 1):
            results = BHSingleSimulator(forwardCurve, fZero, gamma, a, b, c, d, eta, 
                              betas, delta, browniensBH1[i], browniensBH2[i])
            # Simulate
            simulatedCurves.append(results[0])
            deflateurs.append(results[1])
            
        end = time()
        print("")
        print("FORWARD SCENARIO SIMULATION COMPLETE")
        print('-------------------------------')
        print('BH Simulation successfully completed in: '+ str(round((end- start)/60, 2)) + ' minutes')
        
        return([simulatedCurves, deflateurs])
        
    except:
        print('Error encountered. Please ensure BH brownians are in the correct format.')
        sys.exit(1)
        

'''
3) HULL & ROTMAN DISCRETISATION SCHEME
===================================================================================================
'''
def UnitSimulator(forwardCurve, fZero, gamma, a, b, c, d, eta, betas, delta, gaussians, 
                  maxProjectionYear = 50, maxMaturity = 40):
    '''Simulates a single entire forward curve scenario using Python gaussians.
    Inputs
    ========
    forwardCurve = Curve intended to be projected. Select the curve length depending on the maximum maturity intended for projection.
    
    dZ1, dZ2 = Gaussians obtained from the Moodys tool
    
    fZero, gamma, a, b, c, d = parameters for the volatility parametrization
    
    NOTE
    =======
    1) Simulation set by default to 50 years. This is what is used in the internal model
    2) No need to define the simulations
    3) This simulation is not just constrained to Moodys Gaussians, any Gaussians in the correct will work
        -For formatting, please check the BH Browniens.xlsx file in the Inputs folder'''
        
    timeSteps = len(forwardCurve[:maxProjectionYear+maxMaturity]) 
    dt = 1
      
    # Set the forward Curve
    curve = np.array(forwardCurve[:maxMaturity+maxProjectionYear])
    curveAntithetique = np.array(forwardCurve[:maxMaturity+maxProjectionYear])
    qCurve = np.power(curve+delta, 1 - eta)/(1 - eta)
    qCurveAntithetique = np.power(curve+delta, 1 - eta)/(1 - eta)
    
    # f and g vectors 
    fValues = fZero + (1 - fZero)*np.exp(-gamma * np.arange(timeSteps))
    gValues = ((a + b * np.arange(1, timeSteps +1)) * np.exp(-c * np.arange(1, timeSteps +1))) + d
    
    # Product of g and betas (Since they have the same index)
    gBeta1 = gValues * betas[0, :timeSteps]
    gBeta2 = gValues * betas[1, :timeSteps]
    
    terminalCurve = []
    terminalCurveAntithetique = []
    deflateurs = []
    deflateursATQ = []
    gaussiansAntithetique = - copy.deepcopy(gaussians)
       
    # Loop to construct curves
    for t in range(1, timeSteps+1):  
        # Calculate Drift
        mu1 = np.array(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*
                             np.cumsum(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]* np.power(curve + delta, eta)/
                                   (1 + curve)))       
        mu2 = np.array(fValues[t-1]*gBeta2[:len(gBeta2)-t+1] *
                             np.cumsum(fValues[t-1]*gBeta2[:len(gBeta2)-t+1]* np.power(curve + delta, eta)/
                                   (1 + curve)))
        mu = mu1 + mu2
        
        mu1A = np.array(fValues[t-1]*gBeta1[:len(gBeta1)-t+1] *
                             np.cumsum(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]* np.power(curveAntithetique + delta, eta)/
                                   (1 + curveAntithetique)))       
        mu2A = np.array(fValues[t-1]*gBeta2[:len(gBeta2)-t+1] *
                             np.cumsum(fValues[t-1]*gBeta2[:len(gBeta2)-t+1]* np.power(curveAntithetique + delta, eta)/
                                   (1 + curveAntithetique)))
        muA = mu1A + mu2A
        
        fCurveSum = np.power(curve + delta, eta-1)
        fCurveSumA = np.power(curveAntithetique + delta, eta-1)
        
        # Recalculate the qCurve as it should be
        # Construct Curve     
        qCurve = (qCurve
                    + dt * (mu 
                    - 0.5 *eta*  fCurveSum * (np.power(fValues[t-1]*gBeta1[:len(gBeta1)-t+1], 2) + 
                                 np.power(fValues[t-1]*gBeta2[:len(gBeta2)-t+1], 2)))
                    + fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*np.sqrt(dt)*gaussians[t-1, 0] 
                    + fValues[t-1]*gBeta2[:len(gBeta2)-t+1]*np.sqrt(dt)*gaussians[t-1, 1] )
        
        qCurveAntithetique = (qCurveAntithetique
            + dt * (muA 
            - 0.5 *eta* fCurveSumA * (np.power(fValues[t-1]*gBeta1[:len(gBeta1)-t+1], 2) + 
                         np.power(fValues[t-1]*gBeta2[:len(gBeta2)-t+1], 2)))
            + fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*np.sqrt(dt)*gaussiansAntithetique[t-1, 0] 
            + fValues[t-1]*gBeta2[:len(gBeta2)-t+1]*np.sqrt(dt)*gaussiansAntithetique[t-1, 1] )
        
        # At end of each year append the curve to TerminalCurve
        terminalCurve.append(np.power((1-eta)*qCurve, 1/(1-eta)) - delta)
        terminalCurveAntithetique.append(np.power((1-eta)*qCurveAntithetique, 1/(1-eta)) - delta)
        deflateurs.append(1/(1 + (np.power((1-eta)*qCurve[0], 1/(1-eta))) - delta))
        deflateursATQ.append(1/(1+ (np.power((1-eta)*qCurveAntithetique[0], 1/(1-eta)))- delta))
        
        qCurve = qCurve[1:] # Delete 1 year from curve
        qCurveAntithetique = qCurveAntithetique[1:]
        curve = np.power((1-eta)*qCurve, 1/(1-eta)) - delta
        curveAntithetique = np.power((1-eta)*qCurveAntithetique, 1/(1-eta)) - delta 
                    
    return([terminalCurve,
            terminalCurveAntithetique,
            [gaussians[:,0], gaussiansAntithetique[:,0]],
            [gaussians[:,1], gaussiansAntithetique[:,1]],
            deflateurs,
            deflateursATQ])


'''
4) FULL PROJECTOR
===================================================================================================
'''
def fullSimulator(forwardCurve, fZero, gamma, a, b, c, d, eta, betas, delta,
                  simulations = 5000, maxProjectionYear = 50, maxMaturity = 40, 
                  outputGaussians = exportedGaussian1Location):
    '''Loop to simulate forward curve scenarios using Moodys gaussians.
        Based on the UnitSimulator function
    
    Inputs
    ========
    forwardCurve = Curve intended to be projected. Select the curve length depending on the maximum maturity intended for projection.

    
    fZero, gamma, a, b, c, d = parameters for the volatility parametrization
    
    zeroCouponMGTest = Binary value allowing for Martingale Tests to be conducted on the generated Scenarios. 
    
    marketConsistencyTest = Binary value allowing for Swaption Monte Carlo Tests
    
    exportGaussians = exports Gaussians generated in the scenario construction
    
    Outputs
    ========
    1) simulatedCurves & deflateurs = All interest rate scenarios
    2) MGTests = all saved in the Results.xlsx file in the Outputs Folder
    3) Gaussians = all exported to the Gaussians.xlsx file in the Outputs Folder
    4) Zero Coupon scenarios are all saved in the ZCScenarios.csv file in the Outputs folder
    '''
    # Set the seed
    start = time()
    simulatedCurves = []
    gaussians1 = []
    gaussians2 = []
    deflateurs = []
    
    # Use gaussians from sobol sequence
    timeSteps = maxProjectionYear + maxMaturity
    
    # Set optimal seed
    np.random.seed(141) # Optimal seed from 1000 seeds tested
    
    gaussiansFull = np.random.randn(int(simulations/2), 90, 2)
      
    for simu in np.arange(int(simulations/2)):
        results = UnitSimulator(forwardCurve, fZero, gamma, a, b, c, d, eta, betas, delta, 
                                gaussiansFull[simu], 
                                maxProjectionYear, maxMaturity)
        
        # Simulate
        simulatedCurves.append(results[0])
        simulatedCurves.append(results[1])
        gaussians1.append(results[2][0])
        gaussians1.append(results[2][1])
        gaussians2.append(results[3][0])
        gaussians2.append(results[3][1])
        deflateurs.append(results[4])
        deflateurs.append(results[5])
        
    end = time()
    print("")
    print("FORWARD SCENARIO SIMULATION COMPLETE")
    print('-------------------------------')
    print('Simulation successfully completed in '+ str(round((end- start)/60, 2)) + ' minutes')
    
    return([simulatedCurves, deflateurs, gaussians1, gaussians2])
        

'''
4b) SOBOL PROJECTOR
===================================================================================================
'''
def sobolSimulator(forwardCurve, fZero, gamma, a, b, c, d, eta, betas, delta, 
                   maxProjectionYear = 50, maxMaturity = 40):
    '''Loop to simulate forward curve scenarios using Moodys gaussians.
        Based on the UnitSimulator function
    
    Inputs
    ========
    forwardCurve = Curve intended to be projected. Select the curve length depending on the maximum maturity intended for projection.

    
    fZero, gamma, a, b, c, d = parameters for the volatility parametrization
    
    zeroCouponMGTest = Binary value allowing for Martingale Tests to be conducted on the generated Scenarios. 
    
    marketConsistencyTest = Binary value allowing for Swaption Monte Carlo Tests
    
    exportGaussians = exports Gaussians generated in the scenario construction
    
    Outputs
    ========
    1) simulatedCurves & deflateurs = All interest rate scenarios
    2) MGTests = all saved in the Results.xlsx file in the Outputs Folder
    3) Gaussians = all exported to the Gaussians.xlsx file in the Outputs Folder
    4) Zero Coupon scenarios are all saved in the ZCScenarios.csv file in the Outputs folder
    '''
    # Set the seed
    start = time()
    simulatedCurves = []
    gaussians1 = []
    gaussians2 = []
    deflateurs = []

    # Import Sobol gaussians
    gaussiansFull = pd.read_excel(loc + '\Analysis\Sobol Gaussians\SobolScrambled.xlsx').to_numpy()    
    gaussiansFull = [np.array([gaussians[:90], 
                      gaussians[90:]]).T for gaussians in gaussiansFull[:1500]]
    
    for simu in np.arange(1500):
        results = UnitSimulator(forwardCurve, fZero, gamma, a, b, c, d, eta, betas, delta, 
                                gaussiansFull[simu], 
                                maxProjectionYear, maxMaturity)
        
        # Simulate
        simulatedCurves.append(results[0])
        simulatedCurves.append(results[1])
        gaussians1.append(results[2][0])
        gaussians1.append(results[2][1])
        gaussians2.append(results[3][0])
        gaussians2.append(results[3][1])
        deflateurs.append(results[4])
        deflateurs.append(results[5])
        
    end = time()
    print("")
    print("FORWARD SCENARIO SIMULATION COMPLETE")
    print('-------------------------------')
    print('Simulation successfully completed in '+ str(round((end- start)/60, 2)) + ' minutes')
    
    return([simulatedCurves, deflateurs, gaussians1, gaussians2])
 

  
'''
5) DISTRIBUTION VISUALIZER
'''
def visualizeDistributions(simulatedCurves, delta, eta, maxMaturity = 40, maxProjectionYear = 50,
                           ratesDistribLocation = ratesElasticityLocation, zcDistribLocation = zcElasticityLocation):
    '''Calculates the distributions and saves them in an Excel File.
    
    Note that these are Time to Maturity distributions.
    '''
      
    '''OBTAIN THE DISTRIBUTIONS
    ==================================================================================================='''
    # Sortir les prix Zero Coupon  
    zCoupons = copy.deepcopy(simulatedCurves)
    
    # Transform Forwards to ZC
    for i in range(len(simulatedCurves)):
        for j in range(len(simulatedCurves[0])):
            zCoupons[i][j] = np.cumprod(1/(1+simulatedCurves[i][j]))
    
    # Change to Time to Maturity - Select the diagonals
    ZCTimeToMaturity = copy.deepcopy(simulatedCurves)
    for i in range(len(zCoupons)):
        for j in range(maxMaturity): # Select the jth item from each vector ie the diagonals 
            ZCTimeToMaturity[i][j] = [x[j] for x in zCoupons[i][:(len(zCoupons[i][0])-j)]]  #Vector of length maturity + projection - j
    
          
    #Select the boundaries required
    ZCdistributions = np.zeros((maxMaturity*len(ZCTimeToMaturity), maxProjectionYear))
    for i in range(len(ZCTimeToMaturity)):
        for j in range(maxMaturity):
            ZCdistributions[i*maxMaturity+j] = np.array(ZCTimeToMaturity[i][j][:maxProjectionYear])
                
    # Select the maturities we want to analyze
    distributionMaturities = [1, 5, 10, 20, 40]
    quantiles = [0.5, 1, 5, 10, 25, 50, 75, 90, 95, 99, 99.5]
    
    finalZCDistributions = []
    finalRatesDistributions = []
    negativeRatesProportion = []
    
    #Obtain all the distributions for each maturity
    for maturity in distributionMaturities:
        distribs =  ZCdistributions[np.arange(maturity - 1, len(ZCdistributions), maxMaturity)]
        rates = np.array([np.append([EIOPACurve[maturity - 1]], elem) for elem in np.power(1/distribs, 1/maturity) -1])
        finalZCDistributions.append(np.percentile(distribs, quantiles, axis = 0))
        finalRatesDistributions.append(np.percentile(rates, quantiles, axis = 0))
        negativeRatesProportion.append(np.sum(rates.transpose() < 0, axis = 1)/len(simulatedCurves))
        
    # Save Rates Distribution to Excel
    book = load_workbook(ratesDistribLocation)
    writer = pd.ExcelWriter(ratesDistribLocation, engine = 'openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    for i in range(len(finalRatesDistributions)):    
        pd.DataFrame(list(reversed(finalRatesDistributions[i])), 
                 index = list(reversed(quantiles))).to_excel(writer, 
                                        "Rate Distributions "+ str(round(delta, 2)),
                                        startcol = 1,
                                        startrow = i*(len(quantiles)+3)+1,
                                        index = True,
                                        header = True)
   
        pd.DataFrame(negativeRatesProportion[i]).transpose().to_excel(writer, 
                                        "Rate Distributions "+ str(round(delta, 2)),
                                                                   startcol = 1,
                                                                   startrow = i+71,
                                                                   index = False,
                                                                   header = False)
        
    writer.save()
    
    # Save ZeroCoupons Distribution to Excel
    book = load_workbook(zcDistribLocation)
    writer = pd.ExcelWriter(zcDistribLocation, engine = 'openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    for i in range(len(finalZCDistributions)):
        pd.DataFrame(list(reversed(finalZCDistributions[i])), 
                 index = list(reversed(quantiles))).to_excel(writer, 
                                        "ZC Distributions "+ str(round(delta, 2)),
                                        startcol = 1,
                                        startrow = i*(len(quantiles)+3)+1,
                                        index = True,
                                        header = True)
    
    writer.save()

    
'''
6) MONTE CARLO PROJECTOR
===================================================================================================
'''
# Classic Pricer with Python Gaussians
###################################################
def UnitMonteCarloPayoff(forwardCurve, expiry, tenor, fZero, gamma, a, b, c, d, eta, betas, delta, strike):
    '''Simulates a single swaption scenario using Python gaussians.
    Inputs
    ========
    forwardCurve = Curve intended to be projected. Select the curve length depending on the maximum maturity intended for projection.
    
    fZero, gamma, a, b, c, d = parameters for the volatility parametrization
    '''
    forwardCurve = np.array(forwardCurve)[:tenor+expiry-1]
    timeSteps = len(forwardCurve) 
    dt = 1
      
    # Set the forward Curve
    curve = np.array(forwardCurve)
    curveAntithetique = np.array(forwardCurve)
    qCurve = np.power(curve + delta, 1 - eta)/(1 - eta)
    qCurveAntithetique = np.power(curve + delta, 1 - eta)/(1 - eta)
    
    # f and g vectors 
    fValues = fZero + (1 - fZero)*np.exp(- gamma * np.arange(len(forwardCurve)))
    gValues = ((a + b * np.arange(1, len(forwardCurve) +1)) * np.exp(- c * np.arange(1, len(forwardCurve) +1))) + d
    
    # Product of g and betas (Since they have the same index)
    gBeta1 = gValues * betas[0, :tenor+expiry-1]
    gBeta2 = gValues * betas[1, :tenor+expiry-1]

    gaussians = np.random.standard_normal((timeSteps, 2))
    gaussiansAntithetique = - copy.deepcopy(gaussians)
    
    # These will be important in discouting the forward payoff for each scenario
    deflateur = zeroCouponCurve[0]
    deflateurA = zeroCouponCurve[0]
    DF, DFA = 1, 1
       
    # Loop to construct curves
    for t in range(1, expiry+1):
        
        # Calculate Drift
        mu1 = np.array(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*
                             np.cumsum(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]* np.power(curve + delta, eta)/
                                   (1 + curve)))       
        mu2 = np.array(fValues[t-1]*gBeta2[:len(gBeta2)-t+1] *
                             np.cumsum(fValues[t-1]*gBeta2[:len(gBeta2)-t+1]* np.power(curve + delta, eta)/
                                   (1 + curve)))
        mu = mu1 + mu2
        
        mu1A = np.array(fValues[t-1]*gBeta1[:len(gBeta1)-t+1] *
                             np.cumsum(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]* np.power(curveAntithetique + delta, eta)/
                                   (1 + curveAntithetique)))       
        mu2A = np.array(fValues[t-1]*gBeta2[:len(gBeta2)-t+1] *
                             np.cumsum(fValues[t-1]*gBeta2[:len(gBeta2)-t+1]* np.power(curveAntithetique + delta, eta)/
                                   (1 + curveAntithetique)))
        muA = mu1A + mu2A
        
        fCurveSum = np.power(curve + delta, eta-1)
        fCurveSumA = np.power(curveAntithetique + delta, eta-1)
        
        # Recalculate the qCurve as it should be
        # Construct Curve     
        qCurve = (qCurve
                    + dt * (mu 
                    - 0.5 *eta*  fCurveSum * (np.power(fValues[t-1]*gBeta1[:len(gBeta1)-t+1], 2) + 
                                 np.power(fValues[t-1]*gBeta2[:len(gBeta2)-t+1], 2)))
                    + fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*np.sqrt(dt)*gaussians[t-1, 0] 
                    + fValues[t-1]*gBeta2[:len(gBeta2)-t+1]*np.sqrt(dt)*gaussians[t-1, 1] )
        
        qCurveAntithetique = (qCurveAntithetique
            + dt * (muA 
            - 0.5 *eta* fCurveSumA * (np.power(fValues[t-1]*gBeta1[:len(gBeta1)-t+1], 2) + 
                         np.power(fValues[t-1]*gBeta2[:len(gBeta2)-t+1], 2)))
            + fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*np.sqrt(dt)*gaussiansAntithetique[t-1, 0] 
            + fValues[t-1]*gBeta2[:len(gBeta2)-t+1]*np.sqrt(dt)*gaussiansAntithetique[t-1, 1] )
        
        # Find the discount factor at each timestep
        deflateur = deflateur * DF
        DF = 1/(1 + (np.power((1-eta)*qCurve[0], 1/(1-eta))) - delta)
        
        qCurve = qCurve[1:] # Delete 1 year from curve
        curve = np.power((1-eta)*qCurve, 1/(1-eta)) - delta
        
        # Find the discount factor at each timestep
        deflateurA = deflateurA * DFA
        DFA = 1/(1 + (np.power((1-eta)*qCurveAntithetique[0], 1/(1-eta))) - delta)
        
        qCurveAntithetique = qCurveAntithetique[1:] # Delete 1 year from curve
        curveAntithetique = np.power((1-eta)*qCurveAntithetique, 1/(1-eta)) - delta
               
        
    # Move from Forward Curve to Swap Forward Rate 
    ZCCurve = np.append([DF], DF * np.cumprod(1/(1+curve)))
    PVBP = np.sum(ZCCurve)
    forwardSwap = (1 - ZCCurve[len(ZCCurve)-1])/PVBP
    PVBPDeflated  = PVBP * deflateur
    payoff = np.maximum(forwardSwap - strike, 0) * PVBPDeflated
    
    
    ZCCurveAntithetique = np.append([DFA], DFA * np.cumprod(1/(1+curveAntithetique)))
    PVBPAntithetique = np.sum(ZCCurveAntithetique)
    forwardSwapAntithetique = (1 - ZCCurveAntithetique[len(ZCCurveAntithetique)-1])/PVBPAntithetique
    PVBPDeflatedA  = PVBPAntithetique * deflateurA
    payoffAntithetique = np.maximum(forwardSwapAntithetique - strike, 0) * PVBPDeflatedA
    
    payoff = (payoff + payoffAntithetique)/2
    
    return(payoff)

# Classic Monte Carlo with Moodys Gaussians
###################################################
def BHUnitMonteCarloPayoff(forwardCurve, expiry, tenor, fZero, gamma, a, b, c, d, eta, betas, delta, dZ1, dZ2, strike):
    '''Simulates a single entire swaption scenario using Moodys gaussians.
    Inputs
    ========
    forwardCurve = Curve intended to be projected. Select the curve length depending on the maximum maturity intended for projection.
    
    fZero, gamma, a, b, c, d = parameters for the volatility parametrization
    
    dZ1, dZ2 = Gaussians obtained from the Moodys tool
    '''
    forwardCurve = np.array(forwardCurve)[:tenor+expiry-1]
    timeSteps = len(forwardCurve) 
    dt = 1
      
    # Set the forward Curve
    curve = np.array(forwardCurve)
    qCurve = np.power(curve + delta, 1 - eta)/(1 - eta)
    
    # f and g vectors 
    fValues = fZero + (1 - fZero)*np.exp(- gamma * np.arange(len(forwardCurve)))
    gValues = ((a + b * np.arange(1, len(forwardCurve) +1)) * np.exp(- c * np.arange(1, len(forwardCurve) +1))) + d
    
    # Product of g and betas (Since they have the same index)
    gBeta1 = gValues * betas[0, :tenor+expiry-1]
    gBeta2 = gValues * betas[1, :tenor+expiry-1]

    gaussians = np.array([dZ1[:tenor+expiry-1], dZ2[:tenor+expiry-1]]).transpose()
    
    deflateur = zeroCouponCurve[0]
    DF = 1
    
    # Loop to construct curves
    for t in range(1, expiry+1):
        
        # Calculate Drift
        mu1 = np.array(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*
                             np.cumsum(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]* np.power(curve + delta, eta)/
                                   (1 + curve)))       
        mu2 = np.array(fValues[t-1]*gBeta2[:len(gBeta2)-t+1] *
                             np.cumsum(fValues[t-1]*gBeta2[:len(gBeta2)-t+1]* np.power(curve + delta, eta)/
                                   (1 + curve)))
        mu = mu1 + mu2
        
        fCurveSum = np.power(curve + delta, eta-1)
        
        # Recalculate the qCurve as it should be
        # Construct Curve     
        qCurve = (qCurve
                    + dt * (mu 
                    - 0.5 *eta*  fCurveSum * (np.power(fValues[t-1]*gBeta1[:len(gBeta1)-t+1], 2) + 
                                 np.power(fValues[t-1]*gBeta2[:len(gBeta2)-t+1], 2)))
                    + fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*np.sqrt(dt)*gaussians[t-1, 0] 
                    + fValues[t-1]*gBeta2[:len(gBeta2)-t+1]*np.sqrt(dt)*gaussians[t-1, 1] )
              
        # Find the discount factor at each timestep
        deflateur = deflateur * DF
        DF = 1/(1 + (np.power((1-eta)*qCurve[0], 1/(1-eta))) - delta)
        
        qCurve = qCurve[1:] # Delete 1 year from curve
        curve = np.power((1-eta)*qCurve, 1/(1-eta)) - delta
        
    # Move from Forward Curve to Swap Forward Rate 
    meanZCCurve = np.append([DF], DF * np.cumprod(1/(1+curve)))
    meanPVBP = np.sum(meanZCCurve)
    forwardSwap = (1 - meanZCCurve[len(meanZCCurve)-1])/meanPVBP
      
    PVBPDeflated  = meanPVBP * deflateur
    payoff = np.maximum(forwardSwap - strike, 0) * PVBPDeflated
    
    return(payoff)


'''
6) MONTE CARLO PRICER
==================================================================================================='''
def MonteCarloPricer(forwardCurve, expiry, tenor, fZero, gamma, a, b, c, d, eta, betas, delta, simulations):
    '''Loop to simulate swaption price scenarios using Python gaussians.
        Based on the UnitMonteCarloPayoff function
    
    Inputs
    ========
    forwardCurve = Curve intended to be projected. Select the curve length depending on the maximum maturity intended for projection.

    fZero, gamma, a, b, c, d = parameters for the volatility parametrization'''
    start = time()
    payOff= []
    strike = forwardSwapRateCalib(expiry, tenor)
    
    for simu in np.arange(int(simulations/2)):
        payOff.append(UnitMonteCarloPayoff(forwardCurve, expiry, tenor, fZero, gamma, a, b, c, d, eta, 
                                      betas, delta, strike))

    
    price = np.nanmean(payOff) 
    end = time()
    
    timetaken = round((end -start)/60, 2)
    print(str(expiry)+' X '+str(tenor)+ ' Simulation Completed: '+str(timetaken) + ' minutes')
    
    return(price)


# Moody's Monte Carlo Pricer
def MonteCarloPricerBH(forwardCurve, expiry, tenor, fZero, gamma, a, b, c, d, eta, betas, delta, gaussians1, gaussians2):
    '''Loop to simulate swaption price scenarios using Moodys gaussians.
        Based on the BHUnitMonteCarloPayoff function
    
    Inputs
    ========
    forwardCurve = Curve intended to be projected. Select the curve length depending on the maximum maturity intended for projection.
  
    fZero, gamma, a, b, c, d = parameters for the volatility parametrization'''
    start = time()
    payOff= []
    strike = forwardSwapRateCalib(expiry, tenor)
    
    for simu in np.arange(int(len(gaussians1))):
        payOff.append(BHUnitMonteCarloPayoff(forwardCurve, expiry, tenor, fZero, gamma, a, b, c, d, eta, 
                                      betas, delta, gaussians1[simu], gaussians2[simu], strike))

    end = time()    
    price = np.nanmean(payOff)
    
    timetaken = round((end -start)/60, 2)
    print(str(expiry)+' X '+str(tenor)+ ' Simulation Completed: '+str(timetaken) + ' minutes')
    
    return(price)


'''7) GAUSSIAN SEED OPTIMIZER
============================================
'''
def seededSimulator(forwardCurve, fZero, gamma, a, b, c, d, eta, betas, delta,
                  simulations = 5000, maxProjectionYear = 50, maxMaturity = 40):
    errorIndex = []
    
    for k in np.arange(100, 200):
        start = time()
        simulatedCurves = []
        deflateurs = []
    
        # Use gaussians from sobol sequence
        timeSteps = maxProjectionYear + maxMaturity
        np.random.seed(k)
        gaussiansFull = np.random.randn(simulations, 90, 2)
        
        # Save Seed
        #np.save(loc+'\Analysis\Seeds Python\Seed '+str(i)+ '.npy', gaussiansFull)
        
        for simu in np.arange(int(simulations/2)):
            results = UnitSimulator(forwardCurve, fZero, gamma, a, b, c, d, eta, betas, delta, 
                                    gaussiansFull[simu], 
                                    maxProjectionYear, maxMaturity)
            
            # Simulate
            simulatedCurves.append(results[0])
            simulatedCurves.append(results[1])
            deflateurs.append(results[4])
            deflateurs.append(results[5])
        
        '''MARTINGALE TESTS
        ====================================='''
        # Obtain Discount Factors for each scenario
        DF = [np.append(zeroCouponCurve[0],
                        zeroCouponCurve[0]*np.cumprod(df)) for df in deflateurs]
        
        # Average Deflateur
        AvgDeflateur = np.mean(DF, axis = 0) 
        
        # Obtain the Zero Coupon Bonds
        zCoupons = copy.deepcopy(simulatedCurves)
        zCouponsTilde = copy.deepcopy(simulatedCurves)
        
        # Calculate Zero Coupons and Deflated Zero Coupons
        ###################################################################################
        for i in range(len(simulatedCurves)):
            for j in range(len(simulatedCurves[0])):
                zCoupons[i][j] = np.cumprod(1/(1+simulatedCurves[i][j]))
                
                # Multiply each ZC by the discount factor to obtain Discount factors at each timestep
                zCouponsTilde[i][j] = zCoupons[i][j]*DF[i][j]
        
        # Calculate Average Deflated ZC
        ##################################################################################
        # Copy the triangle shape
        resMean = copy.deepcopy(zCoupons[0])
        
        # For calculation of the mean Deflated ZCoupons
        for i in range(len(resMean)):
            resMean[i] = [0]*len(resMean[i])
        
        # Calculate the mean Deflated ZCoupons    
        for i in range(len(zCouponsTilde)):
            for j in range(len(zCouponsTilde[0])):
                resMean[j] = np.nansum([resMean[j], zCouponsTilde[i][j]], axis = 0)
                
        # Final Mean len(simulatedCurves) = number of simulations        
        zCouponTildeAvg = [i/len(simulatedCurves) for i in resMean]
    
        # Average of ZC Simulations
        ##################################################################################
        # Copy the shape of the 1st triangle of the ZCoupons simulations
        zcMean = copy.deepcopy(zCoupons[0])
        
        # For calculation of the mean ZCoupons
        for i in range(len(resMean)):
            zcMean[i] = [0]*len(zcMean[i])
        
        # Calculate the mean ZCoupon len(simulatedCurves) = number of simulations
        for i in range(len(zCouponsTilde)):
            for j in range(len(zCouponsTilde[0])):
                zcMean[j] = np.nansum([zcMean[j], zCoupons[i][j]], axis = 0)
        
        zCouponAvg = [i/len(simulatedCurves) for i in zcMean]
        
        
        'CONVERT TRIANGLE TO RECTANGLE'
        '===================================='
        # Average deflated ZC (Take the First 30 maturities  for each year projected)
        rectangleZCTildeAvg = [zCouponTildeAvg[i][:30] for i in range(50)]
        
        # First Simulation
        ZCSimu1 = [zCoupons[0][i][:30] for i in range(len(zCoupons[0]))]
        forwardsSimu1 = [simulatedCurves[0][i][:30] for i in range(len(simulatedCurves[0]))]
    
        # Error Calculation
        errorsMGTest = [abs((rectangleZCTildeAvg[i]/esgaviva.zeroCouponCurve[1+i:31+i]) -1)
                                    for i in range(len(rectangleZCTildeAvg))]
        
        sqErrors = np.sum(np.power(errorsMGTest, 2))
        errorIndex.append(np.sum(sqErrors))
        end = time()
        print(k)
        
    pd.DataFrame(errorIndex).to_csv(loc + '\Analysis\Seeds Python\ErrorIndex.txt')
    
