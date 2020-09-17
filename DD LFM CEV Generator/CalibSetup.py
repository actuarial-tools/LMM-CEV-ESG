'''
**********************************************************************************************************
*                                        I) CURVE SETUP                                                  *
**********************************************************************************************************
'''
# Extra Packages
from openpyxl import load_workbook
import re
import scipy.optimize as opt
from lmfit import minimize, Parameters, Parameter, report_fit
from itertools import compress
import seaborn as sns
from sklearn.decomposition import PCA
from zipfile import ZipFile
import xlwings as xlw
from datetime import datetime
from nelson_siegel_svensson import NelsonSiegelSvenssonCurve
from nelson_siegel_svensson.calibrate import calibrate_nss_ols

# General management
import numpy as np
from math import *
import pandas as pd
from scipy import stats, optimize
import itertools
import matplotlib.pyplot as plt

# Management of dates and time
from datetime import datetime
from dateutil.relativedelta import relativedelta
from time import strftime, time, localtime
import copy

# Curve Fitting: Might be useful but Python Algorithm is quite weak
from nelson_siegel_svensson.calibrate import calibrate_nss_ols

# Pretty Table: Printing result tables from caalibration
from prettytable import PrettyTable

# Manage operating systems
import os
import sys

# General use packages
from random import seed                # Sets the seed
from openpyxl import load_workbook     # Interfacing with Excel Workbooks

import copy                            # Proper copying of lists
from sklearn.decomposition import PCA

# Change the maximum number of decimal places 
from decimal import *
getcontext().prec = 28 #Ensure calculation to the maximum number of decimal places allowed in Python 


'''R PACKAGE INSTALLATION
=================================================================================================================

Please note that this section may be affected upon version changes of the R package'''
# Import the importr package - allows us to import R packages
from rpy2.robjects.packages import importr

# stats package generates Chi2 CDF for swaption pricing
statsR = importr('stats')

# Import rpy2's package module
import rpy2.robjects.packages as rpackages

# Allow for numpy use in the R function
from rpy2.robjects import numpy2ri
numpy2ri.activate()

utils = rpackages.importr('utils') # import R's utility package. For performing install.packages() R function

# Select a mirror for R packages
utils.chooseCRANmirror(ind=1) # select the first mirror in the list
utils.install_packages('minpack.lm')

# Import Levenberg Marquardt algorithm used in R
minPack = importr('minpack.lm') 


''' CALIBRATION AND PROJECTION INPUTS
==================================================================================================='''
'''EIOPA CURVE
------------------------------------'''
EIOPACurve = np.array(pd.read_excel('Inputs\\EIOPA Curve.xlsx',
              header = None,
             index_col=0))

# Convert Rates to ZC Prices {(1/(1 + rate))^maturity}
EIOPAZCCurve = [pow(1/(1+EIOPACurve[i-1]), i) for i in range(1, len(EIOPACurve)+1)]

# Obtain the vol surface {(ZC(Tk-1)/ZC(Tk))-1}
EIOPAForwardCurve = [(EIOPAZCCurve[i-1]/EIOPAZCCurve[i])-1 for i in range(1, len(EIOPAZCCurve))]

# Define the module ZC and Forward Curves
'Note that this curve will be used for the entirety of the model, to change it, please change the input EIOPA Curve.xlsx in the Inputs Folder'
zeroCouponCurve = np.array(EIOPAZCCurve)
forwardCurve = np.array(EIOPAForwardCurve)


'''Barrie & Hibbert betas'''
betas = pd.read_excel("Inputs\\Betas YE19.xlsx").transpose().to_numpy()



'''
**********************************************************************************************************
*                                       II) CALIBRATION                                                  *
**********************************************************************************************************
'''
'''PRICERS CALIBRATION
========================='''
def forwardSwapRateCalib(expiry, tenor):
    '''Quick Calculation of the forward swap rate based on the module Zero Coupon Curve'''
    num = zeroCouponCurve[expiry-1] - zeroCouponCurve[expiry+tenor-1]
    
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
    return(num/PVBP)

#Black Pricer - Payer
######################################
def blackPayerShifted(N, S0, K, expiry, tenor, sigma, delta):
    '''Calculation of the swaption price based on the Shifted Black Model
    Inputs
    =======
    N - Notional
    S0 - Forward Swap Rate
    K - Strike
    expiry, tenor, sigma are self explanatory
    delta = shift on the Black Model'''     
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
    # Define Black Parameters
    T = expiry
    d1 = (log((S0+delta)/(K + delta)) + 0.5*(pow(sigma,2)*T))/(sigma * sqrt(T))
    d2 = d1 - sigma*sqrt(T)
    
    price = N*PVBP * ((S0+delta)*stats.norm.cdf(d1) - (K+delta)*stats.norm.cdf(d2))
    return(price)


# Normal Pricer - Payer
######################################
def normalPayer(N, S0, K, expiry, tenor, sigma):
    '''Calculation of the payer swaption price based on the Bachelier Model
    Inputs
    =======
    N - Notional
    S0 - Forward Swap Rate
    K - Strike
    expiry, tenor, sigma are self explanatory'''  
    
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
     # Define Black Parameters
    T = expiry
    d1 = (S0 - K)/(sigma * sqrt(T))
    
    #Calculate Price
    price = N * PVBP * ((S0 - K)*stats.norm.cdf(d1) + (sigma * sqrt(T)* exp(-0.5 * pow(d1, 2)))/sqrt(2 * pi))
    return(price)

# Normal Pricer - Receiver
######################################
def normalReceiver(N, S0, K, expiry, tenor, sigma):
    '''Calculation of the receiver swaption price based on the Bachelier Model
    Inputs
    =======
    N - Notional
    S0 - Forward Swap Rate
    K - Strike
    expiry, tenor, sigma are self explanatory'''
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
     # Define Black Parameters
    T = expiry
    d1 = (S0 - K)/(sigma * sqrt(T))
    
    #Calculate Price
    price = N * PVBP * ((K - S0)*stats.norm.cdf(-d1) + (sigma * sqrt(T)* exp(-0.5 * pow(d1, 2)))/sqrt(2 * pi))
    return(price)

# Black Shifted - Payer
######################################
def blackShifted(N, S0, K, expiry, tenor, sigma, delta):
    '''Calculation of the ATM payer swaption price based on the Shifted Black Model
    Inputs
    =======
    N - Notional
    S0 - Forward Swap Rate
    K - Strike
    expiry, tenor, sigma are self explanatory
    delta = Shift on the Black Model'''
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
    # Define Black Parameters
    T = expiry
    d1 = (log((S0+delta)/(K + delta)) + 0.5*(pow(sigma,2)*T))/(sigma * sqrt(T))
    d2 = d1 - sigma*sqrt(T)
    
    price = N * PVBP * ((S0+delta)*stats.norm.cdf(d1) - (K+delta)*stats.norm.cdf(d2))
    return(price)

# Black Shifted - Payer
######################################
def blackShiftedCalib(expiry, tenor, sigma, delta):
    '''Calculation of the ATM payer swaption price based on the Shifted Black Model
    Inputs
    =======
    expiry, tenor, sigma are self explanatory
    delta = shift on the Black Model'''
    S0 = forwardSwapRateCalib(expiry, tenor)
    K = forwardSwapRateCalib(expiry, tenor)
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
    # Define Black Parameters
    T = expiry
    d1 = (log((S0+delta)/(K + delta)) + 0.5*(pow(sigma,2)*T))/(sigma * sqrt(T))
    d2 = d1 - sigma*sqrt(T)
    
    price = PVBP * ((S0+delta)*stats.norm.cdf(d1) - (K+delta)*stats.norm.cdf(d2))
    return(price)


'''IMPLICIT (BACHELIER) VOLATILITY FUNCTIONS
=============================================='''
# Newton Algorithm
def imp_vol_payer_Newton(N, S0, K, expiry, tenor,Price):
    '''Use the classical Newton Algorithm to find the implied Black vol of a payer swaption
    Inputs
    =======
    expiry, tenor, sigma are self explanatory
    delta = shift on the Black Model
    Price = Swaption Price'''
    
    func = lambda sigma: np.power(normalPayer(N, S0, K, expiry, tenor, sigma) - Price, 2.0)
    vol  = optimize.newton(func, 0.1, tol = 1e-5, maxiter=100000)
    return(vol)

# Nelder Mead Algorithm
def imp_vol_payer_NMead(N, S0, K, expiry, tenor,Price):
    '''Use the classical Nelder Mead Algorithm to find the implied Black vol of a payer swaption
    Inputs
    =======
    expiry, tenor, sigma are self explanatory
    delta = shift on the Black Model
    Price = Swaption Price'''
    
    func = lambda sigma: np.power(normalPayer(N, S0, K, expiry, tenor, sigma) - Price, 2.0)
    vol  = optimize.minimize(func, 0.1, method='nelder-mead',
                             options={'xatol': 1e-6, 'disp': False, 'maxiter':100000})
    return(vol.x)

# Toms748 Algorithm
def imp_vol_payer_toms748(N, S0, K, expiry, tenor,Price):
    '''Use the classical Toms748 Algorithm to find the implied Black vol of a payer swaption
    Inputs
    =======
    expiry, tenor, sigma are self explanatory
    delta = shift on the Black Model
    Price = Swaption Price'''
    func = lambda sigma: np.power(normalPayer(N, S0, K, expiry, tenor, sigma) - Price, 1.0)
    vol  = optimize.toms748(func,a = -1, b = 1, xtol = 1e-5)
    return(vol)

def volNormalATMFunction(expiry, tenor, price):
    '''Use the simple ATM function to find the implied Bachelier vol of a payer swaption
    Inputs
    =======
    expiry, tenor are self explanatory
    Price = Swaption Price'''
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
    volNormal = (price /PVBP) * np.sqrt(2 * pi/expiry)
    return (volNormal)


'''VECTORIZING ALL FUNCTIONS
======================================'''
Toms_Vect = np.vectorize(imp_vol_payer_toms748)
Newton_Vect = np.vectorize(imp_vol_payer_Newton)
NMead_Vect = np.vectorize(imp_vol_payer_NMead)
blackVect = np.vectorize(blackShifted)
volNormalVect = np.vectorize(volNormalATMFunction)
blackPayerShiftedVect = np.vectorize(blackPayerShifted)
forwardSwapRateVect = np.vectorize(forwardSwapRateCalib)
normalPayerVect = np.vectorize(normalPayer)
blackVectCalib = np.vectorize(blackShiftedCalib)
volNormalATMFunctionVect = np.vectorize(volNormalATMFunction)


'''VOLATILITY SURFACE
=====================================
- Some of the functions above are required in the construction of the volatility data
- The forwardSwapRateCalib function is used to calculate strikes
- The normalPayerVect function has been used to calculate the normal Prices based on our surface

'''
'''VOLATILITY SURFACE
------------------------------------'''
# Import square/rectangle dataset and transform it to a 3 columns with Expiry, Tenor and the Volatility
volatilitiesLMMPlus = pd.read_excel('Inputs\\Volatility Surface.xlsx', 
                    index_col = 0).unstack().reset_index(name='value')

# Import square/rectangle dataset and transform it to a 3 columns with Expiry, Tenor and the Weight
weightsLMMPlus = pd.read_excel('Inputs\\Weights.xlsx', 
                               index_col = 0).unstack().reset_index(name='value')

# Name the columns
volatilitiesLMMPlus.columns = ['Tenor', 'Expiry', 'Value']
weightsLMMPlus.columns = ['Tenor', 'Expiry', 'Value']

# Unintuitive naming but used for Martingale and Monte Carlo Tests
expiriesBachelier = weightsLMMPlus['Expiry'].to_numpy() 
tenorsBachelier = weightsLMMPlus['Tenor'].to_numpy()

# Calculate the forward Swap Rates for each Expiry X Tenor couple
strikesLMMPlus = forwardSwapRateVect(weightsLMMPlus['Expiry'],weightsLMMPlus['Tenor'] )

# Calculate the Bachelier Price
normalPricesLMMPlus = normalPayerVect(1, strikesLMMPlus, strikesLMMPlus,
                                      weightsLMMPlus['Expiry'],
                                      weightsLMMPlus['Tenor'],
                                      volatilitiesLMMPlus['Value'])  


'''SIGMA ALPHA BETA FUNCTION
=============================
1) PCA parametrization'''
def sigmaAlphaBetaPCA(expiry, tenor, fZero, gamma, a, b, c, d, eta, delta):
    '''Calculates the sigmaAlphaBeta input in the Chi Square pricer
        (Please refer to Andreasen & Andersen (2002)
        
        Inputs
        =========
        expiry, tenor are self explanatory
        eta = Elasticity of the CEV Model
        delta = shift on the model
        
        Inputs from the PCA volatility parametrization:
        (fZero, gamma, a, b, c d )
        
        Please reference the PCA volatility description in the documentation of the DD LFM CEV.'''
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:tenor+expiry])

    # Calculate the forward swap rate
    swapForward = forwardSwapRateCalib(expiry, tenor)

    # Calculate the ZC bond value at T_alpha and T_beta
    zcAlpha = zeroCouponCurve[expiry-1]
    zcBeta = zeroCouponCurve[expiry+tenor-1]

    # Calculate weights
    '''To be confirmed
    ==================='''
    weights = [(swapForward*np.power(forwardCurve[i-1]+ delta, eta)/
                     (np.power(swapForward + delta, eta)*(1 + forwardCurve[i-1])))* 
                    ((zcBeta/(zcAlpha-zcBeta)) + np.sum(zeroCouponCurve[i:expiry+tenor])/PVBP)
                    for i in np.arange(expiry, expiry+tenor)]

    '''VOLATILITY FUNCTIONS
    ========================'''
    # f and g vectors 
    fValues = fZero + (1 - fZero)*np.exp(-gamma * np.arange(tenor + expiry))
    gValues = ((a + b * np.arange(1, tenor + expiry + 1))* 
                np.exp(-c * np.arange(1, tenor + expiry + 1))) + d
    
    # Product of g and betas (Since they have the same index)
    gBeta1 = gValues * betas[0, :tenor+expiry]
    gBeta2 = gValues * betas[1, :tenor+expiry]

    '''INTEGRAL
    ============'''
    sigmaIntegral1 = np.sum([np.square(np.sum(fValues[i]*gBeta1[expiry-i-1:expiry+tenor-i-1]*weights))
                                for i in range(expiry)])
    sigmaIntegral2 = np.sum([np.square(np.sum(fValues[i]*gBeta2[expiry-i-1:expiry+tenor-i-1]*weights))
                                for i in range(expiry)]) 
    
    return(sigmaIntegral1 + sigmaIntegral2)


'''SIGMA ALPHA BETA FUNCTION: Hagan
===================================='''
def sigmaAlphaBetaHagan(expiry, tenor, fZero, gamma, a, b, c, d, eta, delta):
    '''Calculates the sigmaAlphaBeta input in the Hagan Pricer
        (Please refer to Andreasen & Andersen (2002) and Hagan Woodward 
        documentation on CEV approximation(1998))
        
        Inputs
        =========
        expiry, tenor are self explanatory
        eta = Elasticity of the CEV Model
        delta = shift on the model
        
        Inputs from the PCA volatility parametrization:
        (fZero, gamma, a, b, c d )
        
        Please reference the PCA volatility description in the documentation of the DD LFM CEV.'''
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:tenor+expiry])
    
    # Calculate the forward swap rate
    swapForward = forwardSwapRateCalib(expiry, tenor)

    # Calculate the ZC bond value at T_alpha and T_beta
    zcAlpha = zeroCouponCurve[expiry-1]
    zcBeta = zeroCouponCurve[expiry+tenor-1]
    
    # Calculate weights
    weights = [(swapForward*np.power(forwardCurve[i-1]+ delta, eta)/
                 (np.power(swapForward + delta, eta)*(1 + forwardCurve[i-1])))* 
                ((zcBeta/(zcAlpha-zcBeta)) + np.sum(zeroCouponCurve[i:expiry+tenor])/PVBP)
                for i in np.arange(expiry, expiry+tenor)]
    
    '''VOLATILITY FUNCTIONS
    ========================'''
    # f and g vectors 
    fValues = fZero + (1 - fZero)*np.exp(-gamma * np.arange(tenor + expiry)) # Index begins at 0
    gValues = ((a + b * np.arange(1, tenor + expiry + 1))* 
                np.exp(-c * np.arange(1, tenor + expiry + 1))) + d # Index begins at 1
    
    # Product of g and betas (Since they have the same index)
    gBeta1 = gValues * betas[0, :tenor+expiry]
    gBeta2 = gValues * betas[1, :tenor+expiry]

    '''INTEGRAL
    ============'''
    sigmaIntegral1 = np.sum([np.square(np.sum(fValues[i]*gBeta1[expiry-i-1:expiry+tenor-i-1] * weights))
                                for i in range(expiry)])
    sigmaIntegral2 = np.sum([np.square(np.sum(fValues[i]*gBeta2[expiry-i-1:expiry+tenor-i-1] * weights))
                                for i in range(expiry)]) 
 
    return(np.sqrt((sigmaIntegral1 + sigmaIntegral2)/expiry))


''' Hagan Approximation 
=============================='''
def sigma_CEV_Hagan(expiry, tenor, fZero, gamma, a, b, c, d, eta, delta):
    '''Calculates the equivalent Black volatility using the Hagan approximation
    (Please refer to Hagan Woodward documentation on CEV approximation(1998))
    
    Inputs
    =========
    expiry, tenor are self explanatory
    eta = Elasticity of the CEV Model
    delta = shift on the model
    
    Inputs from the PCA volatility parametrization:
    (fZero, gamma, a, b, c d )
    
    Please reference the PCA volatility description in the documentation of the DD LFM CEV.'''
    sigma =  sigmaAlphaBetaHagan(expiry, tenor, fZero, gamma, a, b, c, d, eta, delta)
    Sav = forwardSwapRateCalib(expiry, tenor) + delta
    
    term1 = sigma/np.power(Sav, 1 - eta)
    term4 = (np.power((1 - eta)*sigma,2)*expiry)/(24 * np.power(Sav,2 - 2*eta))
   
    sigma_black = term1 * (1 + term4)
    
    return(sigma_black)


'''CHI SQUARE PAYER WITH PYTHON 
================================='''
def calibrationChiSquarePayerPython(expiry, tenor, fZero, gamma, a, b, c, d, eta, delta):
    '''Calculates the swaption Payer prices using the Andreasen & Andersen swaption pricing formula
    Uses Python to generate chi square CDF values 
    
    (Please refer to Andreasen & Andersen CEV swaption pricing formula (2002))
    
    Inputs
    =========
    expiry, tenor are self explanatory
    eta = Elasticity of the CEV Model
    delta = shift on the model
    
    Inputs from the PCA volatility parametrization:
    (fZero, gamma, a, b, c d )
    
    Please reference the PCA volatility description in the documentation of the DD LFM CEV.
    
    NOTE
    =====
    * Function produces negative values for swaptions with very tenor = 1 and 2'''
    S0 = forwardSwapRateCalib(expiry, tenor)
    K = forwardSwapRateCalib(expiry, tenor)
    
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
    # Parameters for input into Chi Square CDF
    sigmaSquaredT =  sigmaAlphaBetaPCA(expiry, tenor, fZero, gamma, a, b, c, d, eta, delta)
    d = (pow(K+delta,2 - 2*eta))/(pow(1 - eta,2)* sigmaSquaredT)
    b = 1/(1 - eta)
    f = (pow(S0+delta, 2 - 2*eta))/(pow(1 - eta, 2)* sigmaSquaredT)

    # Calculate Price
    price = PVBP * ((S0+delta)*(1 - stats.ncx2.cdf(d, b+2, f)) - (K+delta)*(stats.ncx2.cdf(f, b, d)))
    return(price)


''' Chi Square Pricer - Payer
=============================='''
def calibrationChiSquarePayer(N, S0, K, expiry, tenor, fZero, gamma, a, b, c, d, eta, delta):
    '''Calculates the swaption Payer prices using the Andreasen & Andersen swaption pricing formula
    Uses R to generate chi square CDF values 
    
    (Please refer to Andreasen & Andersen CEV swaption pricing formula (2002))
    
    Inputs
    =========
    expiry, tenor are self explanatory
    eta = Elasticity of the CEV Model
    delta = shift on the model
    
    Inputs from the PCA volatility parametrization:
    (fZero, gamma, a, b, c d )
    
    Please reference the PCA volatility description in the documentation of the DD LFM CEV.'''    
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
    # Parameters for input into Chi Square CDF
    sigmaSquaredT =  sigmaAlphaBetaPCA(expiry, tenor, fZero, gamma, a, b, c, d, eta, delta)
    d = (pow(K+delta,2 - 2*eta))/(pow(1 - eta,2)* sigmaSquaredT)
    b = 1/(1 - eta)
    f = (pow(S0+delta, 2 - 2*eta))/(pow(1 - eta, 2)* sigmaSquaredT)

    # Calculate Price
    price = N * PVBP * ((S0+delta)*(1.0 - statsR.pchisq(d, b+2.0, f)[0]) - (K+delta)*(statsR.pchisq(f, b, d)[0]))
    return(price)


''' Black Pricer - Payer
=============================='''
def calibrationBlackHaganPayer(N, S0, K, expiry, tenor, fZero, gamma, a, b, c, d, eta, delta):    
    '''Calculates the swaption Payer prices using the Hagan approximation 
    (Please refer to Hagan Woodward documentation on CEV approximation(1998))
    
    Inputs
    =========
    expiry, tenor are self explanatory
    eta = Elasticity of the CEV Model
    delta = shift on the model
    
    Inputs from the PCA volatility parametrization:
    (fZero, gamma, a, b, c d )
    
    Please reference the PCA volatility description in the documentation of the DD LFM CEV.'''     
    # Find Zero Coupon Price at each date and calculate PVBP
    PVBP = np.sum(zeroCouponCurve[expiry:expiry+tenor])
    
    sigma = sigma_CEV_Hagan(expiry, tenor, fZero, gamma, a, b, c, d, eta, delta)
    
    # Define Black Parameters
    T = expiry
    d1 = (log((S0+delta)/(K+delta)) + 0.5*(pow(sigma,2)*T))/(sigma * sqrt(T))
    d2 = d1 - sigma*sqrt(T)
    
    price = PVBP * ((S0+delta)*stats.norm.cdf(d1) - (K+delta)*stats.norm.cdf(d2))
    return(price)



'''CALCULATE THE MARKET PRICES
==================================================================='''
# Vectorize Function
calibrationChiSquarePayerVect = np.vectorize(calibrationChiSquarePayer)
calibrationBlackVect = np.vectorize(calibrationBlackHaganPayer)
sigma_CEV_Hagan_Vect = np.vectorize(sigma_CEV_Hagan)
sigmaAlphaBetaHaganVect = np.vectorize(sigmaAlphaBetaHagan)
sigmaAlphaBetaPCAVect = np.vectorize(sigmaAlphaBetaPCA)
calibrationChiSquarePayerPythonVect = np.vectorize(calibrationChiSquarePayerPython)


'''OPTIMIZATION FUNCTIONS (LEVENBERG MARQUARDT IMPLEMENTATION)
=================================================================='''
# Calibration Function (Python Version)
#############################################
def chiSquareCalibratorPython(initialValues = [0.1,0.1,0.1,0.1,0.1,0.1], 
                              lowerBounds = [0,0,0,0,0,0], 
                              upperBounds = [1,1,1,1,1,1], eta = 0.8, delta = 0.1):
    '''Implementation of the Levenberg Marquardt algorithm in Python to find the optimal value 
        based on a given volatility surface.
        
        Function to be minimized:
            Error = (MarketPrices - ModelPrices)/MarketPrices
        
        INPUTS
        ===========
        1) Volatility Surface
            - Obtained from the Excel File in the LMM Volatility Surface.xlsx in the Inputs File
        
        2) EIOPA Curve
            - Obtained from the Excel File in the EIOPA Curve.xlsx in the Inputs File
            
        3) initialValues
            - Initialization values for the algorithms in this order:
                [fZero, gamma, a, b, c, d]
                
            - Default value: [0.1,0.1,0.1,0.1,0.1,0.1]
            
        4) lowerBounds
            -Fix lower limit for the values
            - Default value: [0.0,0.0,0.0,0.0,0.0,0.0]
            
        5) upperBounds
            -Fix upper limit for the values
            - Default value: [1.0,1.0,1.0,1.0,1.0,1.0]
            
        6) eta is the elasticity metaparameter and delta the shift.
    
    NOTE
    =======
    1) Tolerances have all been set to 1e-19 focusing more on precision than runtime
    '''
        
    '''DEFINE THE PARAMETERS 
    =========================='''
    params = Parameters()
    params.add('fZero', value = initialValues[0], min = lowerBounds[0], max = upperBounds[0])
    params.add('gamma',value = initialValues[1], min = lowerBounds[1], max = upperBounds[1])
    params.add('a',value = initialValues[2], min = lowerBounds[2], max = upperBounds[2])
    params.add('b',value = initialValues[3], min = lowerBounds[3], max = upperBounds[3])
    params.add('c', value = initialValues[4], min = lowerBounds[4], max = upperBounds[4])
    params.add('d', value = initialValues[5], min = lowerBounds[5], max = upperBounds[5])
    
    # Objective Function
    objectiveFunctionLMMPlus = lambda test: weightsLMMPlus['Value'].to_numpy() *  ( normalPricesLMMPlus - 
        calibrationChiSquarePayerVect(1.0, strikesLMMPlus, 
                        strikesLMMPlus, 
                        weightsLMMPlus['Expiry'], 
                        weightsLMMPlus['Tenor'], 
                        test['fZero'].value, 
                        test['gamma'].value, 
                        test['a'].value, 
                        test['b'].value,
                        test['c'].value,
                        test['d'].value,
               eta, delta))/normalPricesLMMPlus       
    
   
    '''OPTIMIZE PARAMETERS PYTHON
    =============================='''
    try:
        startTime = time()
        result = minimize(objectiveFunctionLMMPlus, 
                          params, 
                          method = 'leastsq',
                          maxfev = 100000,
                          ftol = 1e-19,
                          xtol = 1e-19,
                          gtol = 1e-19)
        endTime = time()
    except:
        print('')
        print('CALIBRATION ERROR ENCOUNTERED')
        print('----------------------------------')
        print('Possible errors:')
        print('    i)   Incorrect format of inputs (vol surface or weights.')
        print('    ii)  Boundaries & Inputs should be in NUMERIC format.')
        print('    iii) Saturation of algorithm/Impossible to find solution.')
        
    print("**************************************************************************************")
    print('ETA = ' + str(eta) + ' SHIFT = ' + str(delta))
    print('Python optimization Complete in '+ str(round((endTime - startTime)/60, 2)) + ' minutes.')
    
    '''Print Calibration Results'''
    headers = ["SS Error (Python)", "fZero", "Gamma", "a", "b", "c", "d"]
    t2 = PrettyTable()
    t2.add_column(headers[0], [result.chisqr])
    t2.add_column(headers[1], [round(result.params['fZero'].value, 5)])
    t2.add_column(headers[2], [round(result.params['gamma'].value, 5)])
    t2.add_column(headers[3], [round(result.params['a'].value, 5)])
    t2.add_column(headers[4], [round(result.params['b'].value, 5)])
    t2.add_column(headers[5], [round(result.params['c'].value, 5)])
    t2.add_column(headers[6], [round(result.params['d'].value, 5)])
    print(t2)
    
    finalResults = [result.params['fZero'].value, 
            result.params['gamma'].value, 
            result.params['a'].value, 
            result.params['b'].value, 
            result.params['c'].value,
            result.params['d'].value, eta, delta]
    
    
    # Save these results to the Results File   
    try:
        book = load_workbook('Outputs\\Results.xlsx')
        writer = pd.ExcelWriter('Outputs\\Results.xlsx', engine = 'openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        
        pd.DataFrame(finalResults).to_excel(writer, "Calibration",
                                          startcol = 2, 
                                          startrow = 2, 
                                        index = False, 
                                         header = False)
        
        pd.DataFrame(initialValues).to_excel(writer, "Calibration",
                                          startcol = 1, 
                                          startrow = 2, 
                                        index = False, 
                                         header = False)
        
        pd.DataFrame([eta, delta]).to_excel(writer, "Calibration",
                                          startcol = 1, 
                                          startrow = 8, 
                                        index = False, 
                                         header = False)
        
        pd.DataFrame([strftime("%H:%M %d/%m/%Y", localtime())]).to_excel(writer, "Calibration",
                                  startcol = 0, 
                                  startrow = 13, 
                                  index = False, 
                                  header = False)
        writer.save()
    
    except:
      print('')
      print('ERROR ENCOUNTERED')
      print('---------------------')
      print('Please ensure the Results.xlsx file is closed.')
      print('If error persists, please be sure to delete Results.xlsx and initialize a new empty file.')
      print('You could also recover a backup file in the Annex Folder.')
          
        
    return(finalResults)



# Calibration Function (R Version)
#############################################
def chiSquareCalibratorR(initialValues = [0.1,0.1,0.1,0.1,0.1,0.1], 
                              lowerBounds = [0.0,0.0,0.0,0.0,0.0,0.0], 
                              upperBounds = [1.0,1.0,1.0,1.0,1.0,1.0], eta = 0.8, delta = 0.1):  
    '''Implementation of the Levenberg Marquardt algorithm in R to find the optimal value 
        based on a given volatility surface.
                
        Function to be minimized:
            Error = (MarketPrices - ModelPrices)/MarketPrices
        
        INPUTS
        ===========
        1) Volatility Surface
            - Obtained from the Excel File in the LMM Volatility Surface.xlsx in the Inputs File
        
        2) EIOPA Curve
            - Obtained from the Excel File in the EIOPA Curve.xlsx in the Inputs File
            
        3) initialValues
            - Initialization values for the algorithms in this order:
                [fZero, gamma, a, b, c, d]
        
        4) lowerBounds
            -Fix lower limit for the values
            
        5) upperBounds
            -Fix upper limit for the values
            
        6) eta is the elasticity metaparameter and delta the shift.
    
    NOTE
    =======
    1) Tolerances have all been set to 1e-12 focusing more on precision than runtime
    '''
    
    '''DEFINE THE OBJECTIVE FUNCTION
    ================================='''
    objectiveFunctionR = lambda test:  weightsLMMPlus['Value'].to_numpy() * ( normalPricesLMMPlus - 
    calibrationChiSquarePayerVect(1.0, strikesLMMPlus, 
                    strikesLMMPlus, 
                    weightsLMMPlus['Expiry'].to_numpy(), 
                    weightsLMMPlus['Tenor'].to_numpy(), 
                        test[0], 
                        test[1], 
                        test[2], 
                        test[3],
                        test[4],
                        test[5],
                        eta, delta))/normalPricesLMMPlus 
    
    '''OPTIMIZE PARAMETERS R
    =============================='''
    try:
        startTime = time()
        result = minPack.nls_lm(par= list(initialValues),
                            lower = np.array(lowerBounds),
                            upper = np.array(upperBounds),
                             fn = objectiveFunctionR,
                             control = minPack.nls_lm_control(ftol = 1e-19, 
                                                              gtol = 1e-19,
                                                              ptol = 1e-19))
        endTime = time()
    except:
        print('')
        print('CALIBRATION ERROR ENCOUNTERED')
        print('----------------------------------')
        print('Possible errors:')
        print('    i)   Incorrect format of inputs (vol surface or weights.')
        print('    ii)  Boundaries & Inputs should be in NUMERIC format.')
        print('    iii) Saturation of algorithm/Impossible to find solution.')
    

    
    print(" ")
    print('R optimization Complete in '+ str(round((endTime - startTime)/60, 2)) + ' minutes.')
    
    '''Print Calibration Results'''
    headers = ["SS Error (R)", "fZero", "Gamma", "a", "b", "c", "d"]
    t3 = PrettyTable()
    t3.add_column(headers[0], [result[8][0]])
    t3.add_column(headers[1], [round(result[0][0][0], 5)])
    t3.add_column(headers[2], [round(result[0][1][0], 5)])
    t3.add_column(headers[3], [round(result[0][2][0], 5)])
    t3.add_column(headers[4], [round(result[0][3][0], 5)])
    t3.add_column(headers[5], [round(result[0][4][0], 5)])
    t3.add_column(headers[6], [round(result[0][5][0], 5)])
    print(t3)
    print(" ")
    print("**************************************************************************************")
    print(" ")
    
    finalResults = [result[0][0][0], 
        result[0][1][0], 
        result[0][2][0], 
        result[0][3][0], 
        result[0][4][0],
        result[0][5][0], eta, delta]
    
    
    # Save these results to the Results File
    try:
        book = load_workbook('Outputs\\Results.xlsx')
        writer = pd.ExcelWriter('Outputs\\Results.xlsx', engine = 'openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        
        pd.DataFrame(finalResults).to_excel(writer, "Calibration",
                                          startcol = 3, 
                                          startrow = 2, 
                                        index = False, 
                                         header = False)
    
        pd.DataFrame([strftime("%H:%M %d/%m/%Y", localtime())]).to_excel(writer, "Calibration",
                                      startcol = 0, 
                                      startrow = 16, 
                                      index = False, 
                                      header = False)
        writer.save()
        
    except:
      print('')
      print('ERROR ENCOUNTERED')
      print('---------------------')
      print('Please ensure the Results.xlsx file is closed.')
      print('If error persists, please be sure to delete Results.xlsx and initialize a new empty file.')
      print('You could also recover a backup file in the Annex Folder.')
             
    return(finalResults)


def fullCalibrator(initialValues = [0.1,0.1,0.1,0.1,0.1,0.1], 
                              lowerBounds = [0.0,0.0,0.0,0.0,0.0,0.0], 
                              upperBounds = [1.0,1.0,1.0,1.0,1.0,1.0], eta = 0.8, delta = 0.1):
    
    # Run the Calibrator using the Python minimize Package
    finalResultsPython = chiSquareCalibratorPython(initialValues, lowerBounds, upperBounds, 
                                                   eta, delta)
    
    # Run the Calibrator using the R minpack Package
    finalResultsR = chiSquareCalibratorR(initialValues, lowerBounds, upperBounds, 
                                                   eta, delta)
    
    # Save Results
    return([finalResultsPython, finalResultsR])
'''
**********************************************************************************************************
*                                       III) PROJECTION                                                  *
**********************************************************************************************************
'''

'''
1) BH INPUT BROWNIENS
===================================================================================================
'''
def BHSingleSimulator(forwardCurve, fZero, gamma, a, b, c, d, eta, betas, delta, dZ1, dZ2, 
                      maxProjectionYear = 50, maxMaturity = 40):
    '''Simulates a single entire forward curve scenario using Moodys gaussians.
    
    Inputs
    ========
    forwardCurve = Curve intended to be projected. Select the curve length depending on the maximum maturity intended for projection.
    
    dZ1, dZ2 = Gaussians obtained from the Moodys tool
    
    fZero, gamma, a, b, c, d = parameters for the volatility parametrization
    
    NOTE
    =======
    1) Simulation set by default to 50 years. This is what is used in the internal model
    2) No need to define the simulations
    3) This simulation is not just constrained to Moodys Gaussians, any Gaussians in the correct will work
        -For formatting, please check the BH Browniens.xlsx file in the Inputs folder
    '''
    timeSteps = len(forwardCurve[:maxMaturity+maxProjectionYear]) 
    dt = 1
      
    # Set the forward Curve
    curve = np.array(forwardCurve[:maxMaturity+maxProjectionYear])
    qCurve = np.power(curve+delta, 1 - eta)/(1 - eta)
    
    # f and g vectors 
    fValues = fZero + (1 - fZero)*np.exp(-gamma * np.arange(timeSteps))
    gValues = ((a + b * np.arange(1, timeSteps+1)) * np.exp(-c * np.arange(1, timeSteps+1))) + d
    
    # Product of g and betas (Since they have the same index)
    gBeta1 = gValues * betas[0, :timeSteps]
    gBeta2 = gValues * betas[1, :timeSteps]
    
    terminalCurve = []
    deflateurs = []
    gaussians = np.array([dZ1, dZ2]).transpose()

    
    # Loop to construct curves
    for t in range(1, maxProjectionYear+1):  
        # Calculate Drift
        mu1 = np.array(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*
                             np.cumsum(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]* np.power(curve + delta, eta)/
                                   (1 + curve)))       
        mu2 = np.array(fValues[t-1]*gBeta2[:len(gBeta2)-t+1] *
                             np.cumsum(fValues[t-1]*gBeta2[:len(gBeta2)-t+1]* np.power(curve + delta, eta)/
                                   (1 + curve)))
        mu = mu1 + mu2

        fCurveSum = np.power(curve + delta, eta-1)
        
        # Recalculate the qCurve as it should be
        # Construct Curve     
        qCurve = (qCurve
                    + dt * (mu 
                    - 0.5 *eta*  fCurveSum * (np.power(fValues[t-1]*gBeta1[:len(gBeta1)-t+1], 2) + 
                                 np.power(fValues[t-1]*gBeta2[:len(gBeta2)-t+1], 2)))
                    + fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*np.sqrt(dt)*gaussians[t-1, 0] 
                    + fValues[t-1]*gBeta2[:len(gBeta2)-t+1]*np.sqrt(dt)*gaussians[t-1, 1] )
        
        # Since the timestep is 1 we do this for each simulation
        terminalCurve.append(np.power((1-eta)*qCurve, 1/(1-eta)) - delta)
        deflateurs.append(1/(1 + (np.power((1-eta)*qCurve[0], 1/(1-eta))) - delta))
        
        qCurve = qCurve[1:] # Delete 1 year from curve
        curve = np.power((1-eta)*qCurve, 1/(1-eta)) - delta
                    
    return([terminalCurve,
            deflateurs])


'''
2) DEFINING THE FINAL PROJECTOR FUNCTION
===================================================================================================
'''
def BHfullSimulator(forwardCurve, fZero, gamma, a, b, c, d, eta, betas, delta, 
                    maxProjectionYear = 50, maxMaturity = 40,
                    zeroCouponMGTest = False, SwaptionMonteCarloTest= False, viewDistributions = True):
    '''Loop to simulate forward curve scenarios using Moodys gaussians.
        Based on the BHSingleSimulator function
    
    Inputs
    ========
    forwardCurve = Curve intended to be projected. Select the curve length depending on the maximum maturity intended for projection.
    
    fZero, gamma, a, b, c, d = parameters for the volatility parametrization
    
    zeroCouponMGTest = Binary value allowing for Martingale Tests to be conducted on the generated Scenarios. 
    
    SwaptionMonteCarloTest = Binary value allowing for Swaption Monte Carlo Tests
      
    Outputs
    ========
    1) simulatedCurves & deflateurs = All interest rate scenarios
    2) MGTests = all saved in the Results.xlsx file in the Outputs Folder
    3) Zero Coupon scenarios are all saved in the ZCScenarios.csv file in the Outputs folder
    
    NOTE
    =======
    1) Simulation set by default to 50 years. This is what is used in the internal model
    2) No need to define the simulations
    3) This simulation is not just constrained to Moodys Gaussians, any Gaussians in the correct will work
        -For formatting, please check the BH Browniens.xlsx file in the Inputs folder
    '''
    try:
        start = time()
        # Import data from BH Brownians
        browniansBH = pd.read_excel("Inputs\\Gaussians.xlsx", 
                                    names = ['Trial', 'Timestep', 'Gaussian1', 'Gaussian2'], 
                                    index_col = [0,1])
        maxCol = np.amax(np.array(browniansBH.index.get_level_values(1)))
        simulations = int(len(browniansBH)/maxCol)
        browniensBH1 = browniansBH['Gaussian1'].to_numpy().reshape((simulations, maxCol))
        browniensBH2 = browniansBH['Gaussian2'].to_numpy().reshape((simulations, maxCol))
        
        # Where results will be saved
        simulatedCurves = []
        deflateurs = []
        
        for i in np.arange(simulations, step = 1):
            results = BHSingleSimulator(forwardCurve, fZero, gamma, a, b, c, d, eta, 
                              betas, delta, browniensBH1[i], browniensBH2[i])
            # Simulate
            simulatedCurves.append(results[0])
            deflateurs.append(results[1])
            
        end = time()
        print("")
        print("FORWARD SCENARIO SIMULATION COMPLETE")
        print('-------------------------------')
        print('BH Simulation successfully completed in: '+ str(round((end- start)/60, 2)) + ' minutes')
    
    except:
        print('Error encountered. Please ensure BH brownians are in the correct format.')
        sys.exit(1)
        
    if (zeroCouponMGTest == True):
      MGTest(simulatedCurves, deflateurs)  
    
    if (SwaptionMonteCarloTest == True):
        print("")
        print("MONTE CARLO SIMULATION")
        print('-------------------------------')
        fullMarketConsistencyTestMoodys(forwardCurve, 
                                        fZero, gamma, a, b, c, d, eta, betas, delta, 
                                        browniensBH1,browniensBH2)        
    if (viewDistributions == True):
       visualizeDistributions(simulatedCurves, delta, eta, maxMaturity = 40, maxProjectionYear = 50)    

    '''SAVE ZERO COUPON SCENARIOS
    ==================================================================================================='''
    # Sortir les prix Zero Coupon  
    zCoupons = copy.deepcopy(simulatedCurves)
    
    # Transform Forwards to ZC
    for i in range(len(simulatedCurves)):
        for j in range(len(simulatedCurves[0])):
            zCoupons[i][j] = np.cumprod(1/(1+simulatedCurves[i][j]))
    
    # Change to Time to Maturity - Select the diagonals
    ZCTimeToMaturity = copy.deepcopy(simulatedCurves)
    for i in range(len(zCoupons)):
        for j in range(maxMaturity): # Select the jth item from each vector ie the diagonals 
            ZCTimeToMaturity[i][j] = [x[j] for x in zCoupons[i][:(len(zCoupons[i][0])-j)]]  #Vector of length maturity + projection - j
    
          
    #Select the boundaries required
    ZCdistributions = np.zeros((maxMaturity*len(ZCTimeToMaturity), maxProjectionYear))
    for i in range(len(ZCTimeToMaturity)):
        for j in range(maxMaturity):
            ZCdistributions[i*maxMaturity+j] = np.array(ZCTimeToMaturity[i][j][:maxProjectionYear])
    
    # Save ZC in csv Format
    index1 = np.array(list(range(1, maxMaturity+1))*simulations) # Set the First index
    index2 = np.repeat(np.arange(1, simulations+1), maxMaturity) # Will repeat the simulation number 40 times
    index2bis = np.array(['Simulation' + str(i) for i in index2 ])
    pd.DataFrame(ZCdistributions, 
                 columns = np.arange(1, maxProjectionYear+1),
                 index = [index2bis, index1]).to_csv('Outputs\\ZCScenarios.csv')
    
    finalEnd = time()
    print('')
    print('FULL SIMULATION COMPLETE')
    print('-------------------------------')
    print('The results of the simulation alongside the tests can be found in the Outputs repertoire.')
    print('The concerned files are: Results.xlsx, Gaussians.xlsx and the Distribution Folder')  
    print('Time log: ' +str(round((finalEnd- start)/60, 2)) + ' minutes')
    
    return([simulatedCurves, deflateurs])


'''
3) HULL & WHITE DISCRETISATION SCHEME
===================================================================================================
'''
def UnitSimulator(forwardCurve, fZero, gamma, a, b, c, d, eta, betas, delta,
                  maxProjectionYear = 50, maxMaturity = 40):
    '''Simulates a single entire forward curve scenario using Python gaussians.
    Inputs
    ========
    forwardCurve = Curve intended to be projected. Select the curve length depending on the maximum maturity intended for projection.
    
    dZ1, dZ2 = Gaussians obtained from the Moodys tool
    
    fZero, gamma, a, b, c, d = parameters for the volatility parametrization
    
    NOTE
    =======
    1) Simulation set by default to 50 years. This is what is used in the internal model
    2) No need to define the simulations
    3) This simulation is not just constrained to Moodys Gaussians, any Gaussians in the correct will work
        -For formatting, please check the BH Browniens.xlsx file in the Inputs folder'''
        
    timeSteps = len(forwardCurve[:maxProjectionYear+maxMaturity]) 
    dt = 1
      
    # Set the forward Curve
    curve = np.array(forwardCurve[:maxMaturity+maxProjectionYear])
    curveAntithetique = np.array(forwardCurve[:maxMaturity+maxProjectionYear])
    qCurve = np.power(curve+delta, 1 - eta)/(1 - eta)
    qCurveAntithetique = np.power(curve+delta, 1 - eta)/(1 - eta)
    
    # f and g vectors 
    fValues = fZero + (1 - fZero)*np.exp(-gamma * np.arange(timeSteps))
    gValues = ((a + b * np.arange(1, timeSteps +1)) * np.exp(-c * np.arange(1, timeSteps +1))) + d
    
    # Product of g and betas (Since they have the same index)
    gBeta1 = gValues * betas[0, :timeSteps]
    gBeta2 = gValues * betas[1, :timeSteps]
    
    terminalCurve = []
    terminalCurveAntithetique = []
    gaussians = []
    deflateurs = []
    deflateursATQ = []
    gaussians = np.random.standard_normal((timeSteps, 2))
    gaussiansAntithetique = - copy.deepcopy(gaussians)
       
    # Loop to construct curves
    for t in range(1, timeSteps+1):  
        # Calculate Drift
        mu1 = np.array(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*
                             np.cumsum(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]* np.power(curve + delta, eta)/
                                   (1 + curve)))       
        mu2 = np.array(fValues[t-1]*gBeta2[:len(gBeta2)-t+1] *
                             np.cumsum(fValues[t-1]*gBeta2[:len(gBeta2)-t+1]* np.power(curve + delta, eta)/
                                   (1 + curve)))
        mu = mu1 + mu2
        
        mu1A = np.array(fValues[t-1]*gBeta1[:len(gBeta1)-t+1] *
                             np.cumsum(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]* np.power(curveAntithetique + delta, eta)/
                                   (1 + curveAntithetique)))       
        mu2A = np.array(fValues[t-1]*gBeta2[:len(gBeta2)-t+1] *
                             np.cumsum(fValues[t-1]*gBeta2[:len(gBeta2)-t+1]* np.power(curveAntithetique + delta, eta)/
                                   (1 + curveAntithetique)))
        muA = mu1A + mu2A
        
        fCurveSum = np.power(curve + delta, eta-1)
        fCurveSumA = np.power(curveAntithetique + delta, eta-1)
        
        # Recalculate the qCurve as it should be
        # Construct Curve     
        qCurve = (qCurve
                    + dt * (mu 
                    - 0.5 *eta*  fCurveSum * (np.power(fValues[t-1]*gBeta1[:len(gBeta1)-t+1], 2) + 
                                 np.power(fValues[t-1]*gBeta2[:len(gBeta2)-t+1], 2)))
                    + fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*np.sqrt(dt)*gaussians[t-1, 0] 
                    + fValues[t-1]*gBeta2[:len(gBeta2)-t+1]*np.sqrt(dt)*gaussians[t-1, 1] )
        
        qCurveAntithetique = (qCurveAntithetique
            + dt * (muA 
            - 0.5 *eta* fCurveSumA * (np.power(fValues[t-1]*gBeta1[:len(gBeta1)-t+1], 2) + 
                         np.power(fValues[t-1]*gBeta2[:len(gBeta2)-t+1], 2)))
            + fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*np.sqrt(dt)*gaussiansAntithetique[t-1, 0] 
            + fValues[t-1]*gBeta2[:len(gBeta2)-t+1]*np.sqrt(dt)*gaussiansAntithetique[t-1, 1] )
        
        # At end of each year append the curve to TerminalCurve
        terminalCurve.append(np.power((1-eta)*qCurve, 1/(1-eta)) - delta)
        terminalCurveAntithetique.append(np.power((1-eta)*qCurveAntithetique, 1/(1-eta)) - delta)
        deflateurs.append(1/(1 + (np.power((1-eta)*qCurve[0], 1/(1-eta))) - delta))
        deflateursATQ.append(1/(1+ (np.power((1-eta)*qCurveAntithetique[0], 1/(1-eta)))- delta))
        
        qCurve = qCurve[1:] # Delete 1 year from curve
        qCurveAntithetique = qCurveAntithetique[1:]
        curve = np.power((1-eta)*qCurve, 1/(1-eta)) - delta
        curveAntithetique = np.power((1-eta)*qCurveAntithetique, 1/(1-eta)) - delta 
                    
    return([terminalCurve,
            terminalCurveAntithetique,
            [gaussians[:,0], gaussiansAntithetique[:,0]],
            [gaussians[:,1], gaussiansAntithetique[:,1]],
            deflateurs,
            deflateursATQ])


'''
4) FULL PROJECTOR
===================================================================================================
'''
def fullSimulator(forwardCurve, fZero, gamma, a, b, c, d, eta, betas, delta,
                  simulations = 5000, maxProjectionYear = 50, maxMaturity = 40,
                  zeroCouponMGTest = False, marketConsistencyTest = False, exportGaussians = False,
                  viewDistributions = True):
    '''Loop to simulate forward curve scenarios using Moodys gaussians.
        Based on the UnitSimulator function
    
    Inputs
    ========
    forwardCurve = Curve intended to be projected. Select the curve length depending on the maximum maturity intended for projection.

    
    fZero, gamma, a, b, c, d = parameters for the volatility parametrization
    
    zeroCouponMGTest = Binary value allowing for Martingale Tests to be conducted on the generated Scenarios. 
    
    marketConsistencyTest = Binary value allowing for Swaption Monte Carlo Tests
    
    exportGaussians = exports Gaussians generated in the scenario construction
    
    Outputs
    ========
    1) simulatedCurves & deflateurs = All interest rate scenarios
    2) MGTests = all saved in the Results.xlsx file in the Outputs Folder
    3) Gaussians = all exported to the Gaussians.xlsx file in the Outputs Folder
    4) Zero Coupon scenarios are all saved in the ZCScenarios.csv file in the Outputs folder
    '''
    # Set the seed
    start = time()
    simulatedCurves = []
    gaussians1 = []
    gaussians2 = []
    deflateurs = []
    
    for simu in np.arange(int(simulations/2)):
        np.random.seed(simu)
        results = UnitSimulator(forwardCurve, fZero, gamma, a, b, c, d, eta, betas, delta,
                                maxProjectionYear, maxMaturity)
        
        # Simulate
        simulatedCurves.append(results[0])
        simulatedCurves.append(results[1])
        gaussians1.append(results[2])
        gaussians2.append(results[3])
        deflateurs.append(results[4])
        deflateurs.append(results[5])
        
    end = time()
    print("")
    print("FORWARD SCENARIO SIMULATION COMPLETE")
    print('-------------------------------')
    print('Simulation successfully completed in '+ str(round((end- start)/60, 2)) + ' minutes')

        
    if (zeroCouponMGTest == True):
      MGTest(simulatedCurves, deflateurs)  
    
    if (marketConsistencyTest == True):
        print("")
        print("MONTE CARLO PRICING")
        print('-------------------------------')
        fullMarketConsistencyTest(forwardCurve, fZero, gamma, a, b, c, d, eta, betas, delta, simulations) 
        print('COMPLETED')
    
    if (exportGaussians == True):
        pd.DataFrame(gaussians1).to_excel('Outputs\\Gaussians1.xlsx',
                                          sheet_name = 'Gaussians1')
        pd.DataFrame(gaussians2).to_excel('Outputs\\Gaussians2.xlsx',
                                          sheet_name = 'Gaussians2')
        
    if (viewDistributions == True):
       visualizeDistributions(simulatedCurves, delta, eta, maxMaturity = 40, maxProjectionYear = 50) 
    
    '''SAVE ZERO COUPON SCENARIOS
    ==================================================================================================='''
    # Sortir les prix Zero Coupon  
    zCoupons = copy.deepcopy(simulatedCurves)
    
    # Transform Forwards to ZC
    for i in range(len(simulatedCurves)):
        for j in range(len(simulatedCurves[0])):
            zCoupons[i][j] = np.cumprod(1/(1+simulatedCurves[i][j]))
    
    # Change to Time to Maturity - Select the diagonals
    ZCTimeToMaturity = copy.deepcopy(simulatedCurves)
    for i in range(len(zCoupons)):
        for j in range(maxMaturity): # Select the jth item from each vector ie the diagonals 
            ZCTimeToMaturity[i][j] = [x[j] for x in zCoupons[i][:(len(zCoupons[i][0])-j)]]  #Vector of length maturity + projection - j
            
    #Select the boundaries required
    ZCdistributions = np.zeros((maxMaturity*len(ZCTimeToMaturity), maxProjectionYear))
    for i in range(len(ZCTimeToMaturity)):
        for j in range(maxMaturity):
            ZCdistributions[i*maxMaturity+j] = np.array(ZCTimeToMaturity[i][j][:maxProjectionYear])
    
    # Save ZC in csv Format
    index1 = np.array(list(range(1, maxMaturity+1))*simulations) # Set the First index
    index2 = np.repeat(np.arange(1, simulations+1), maxMaturity) # Will repeat the simulation number 40 times
    index2bis = np.array(['Simulation' + str(i) for i in index2 ])
    pd.DataFrame(ZCdistributions, 
                 columns = np.arange(1, maxProjectionYear+1),
                 index = [index2bis, index1]).to_csv('Outputs\\ZCScenarios.csv')
            
    finalEnd = time()
    print('')
    print('FULL SIMULATION COMPLETE')
    print('-------------------------------')
    print('The results of the simulation alongside the tests can be found in the Outputs repertoire.')
    print('The concerned files are: ZCScenarios.csv, Results.xlsx, Gaussians.xlsx and the Distribution Folder')
    print('Time log: ' +str(round((finalEnd- start)/60, 2)) + ' minutes')

'''
5) DISTRIBUTION VISUALIZER
'''
def visualizeDistributions(simulatedCurves, delta, eta, maxMaturity = 40, maxProjectionYear = 50):
    '''Calculates the distributions and saves them in an Excel File.
    
    Note that these are Time to Maturity distributions.
    '''
      
    '''OBTAIN THE DISTRIBUTIONS
    ==================================================================================================='''
    # Sortir les prix Zero Coupon  
    zCoupons = copy.deepcopy(simulatedCurves)
    
    # Transform Forwards to ZC
    for i in range(len(simulatedCurves)):
        for j in range(len(simulatedCurves[0])):
            zCoupons[i][j] = np.cumprod(1/(1+simulatedCurves[i][j]))
    
    # Change to Time to Maturity - Select the diagonals
    ZCTimeToMaturity = copy.deepcopy(simulatedCurves)
    for i in range(len(zCoupons)):
        for j in range(maxMaturity): # Select the jth item from each vector ie the diagonals 
            ZCTimeToMaturity[i][j] = [x[j] for x in zCoupons[i][:(len(zCoupons[i][0])-j)]]  #Vector of length maturity + projection - j
    
          
    #Select the boundaries required
    ZCdistributions = np.zeros((maxMaturity*len(ZCTimeToMaturity), maxProjectionYear))
    for i in range(len(ZCTimeToMaturity)):
        for j in range(maxMaturity):
            ZCdistributions[i*maxMaturity+j] = np.array(ZCTimeToMaturity[i][j][:maxProjectionYear])
                
    # Select the maturities we want to analyze
    distributionMaturities = [1, 5, 10, 20, 40]
    quantiles = [0.5, 1, 5, 10, 25, 50, 75, 90, 95, 99, 99.5]
    
    finalZCDistributions = []
    finalRatesDistributions = []
    negativeRatesProportion = []
    
    #Obtain all the distributions for each maturity
    for maturity in distributionMaturities:
        distribs =  ZCdistributions[np.arange(maturity - 1, len(ZCdistributions), maxMaturity)]
        rates = np.array([np.append([EIOPACurve[maturity - 1]], elem) for elem in np.power(1/distribs, 1/maturity) -1])
        finalZCDistributions.append(np.percentile(distribs, quantiles, axis = 0))
        finalRatesDistributions.append(np.percentile(rates, quantiles, axis = 0))
        negativeRatesProportion.append(np.sum(rates.transpose() < 0, axis = 1)/len(simulatedCurves))
        
    # Save Rates Distribution to Excel
    book = load_workbook('Outputs\\Distributions\\Rates Elasticity = '+str(eta)+'.xlsx')
    writer = pd.ExcelWriter('Outputs\\Distributions\\Rates Elasticity = '+str(eta)+'.xlsx', engine = 'openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    for i in range(len(finalRatesDistributions)):    
        pd.DataFrame(list(reversed(finalRatesDistributions[i])), 
                 index = list(reversed(quantiles))).to_excel(writer, 
                                        "Rate Distributions "+ str(round(delta, 2)),
                                        startcol = 1,
                                        startrow = i*(len(quantiles)+3)+1,
                                        index = True,
                                        header = True)
   
        pd.DataFrame(negativeRatesProportion[i]).transpose().to_excel(writer, 
                                        "Rate Distributions "+ str(round(delta, 2)),
                                                                   startcol = 1,
                                                                   startrow = i+71,
                                                                   index = False,
                                                                   header = False)
        
    writer.save()
    
    # Save ZeroCoupons Distribution to Excel
    book = load_workbook('Outputs\\Distributions\\ZeroCoupons Elasticity = '+str(eta)+'.xlsx')
    writer = pd.ExcelWriter('Outputs\\Distributions\\ZeroCoupons Elasticity = '+str(eta)+'.xlsx', engine = 'openpyxl')
    writer.book = book
    writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
    
    for i in range(len(finalZCDistributions)):
        pd.DataFrame(list(reversed(finalZCDistributions[i])), 
                 index = list(reversed(quantiles))).to_excel(writer, 
                                        "ZC Distributions "+ str(round(delta, 2)),
                                        startcol = 1,
                                        startrow = i*(len(quantiles)+3)+1,
                                        index = True,
                                        header = True)
    
    writer.save()

    
'''
6) MONTE CARLO PROJECTOR
===================================================================================================
'''
# Classic Pricer with Python Gaussians
###################################################
def UnitMonteCarloPayoff(forwardCurve, expiry, tenor, fZero, gamma, a, b, c, d, eta, betas, delta, strike):
    '''Simulates a single swaption scenario using Python gaussians.
    Inputs
    ========
    forwardCurve = Curve intended to be projected. Select the curve length depending on the maximum maturity intended for projection.
    
    fZero, gamma, a, b, c, d = parameters for the volatility parametrization
    '''
    forwardCurve = np.array(forwardCurve)[:tenor+expiry-1]
    timeSteps = len(forwardCurve) 
    dt = 1
      
    # Set the forward Curve
    curve = np.array(forwardCurve)
    curveAntithetique = np.array(forwardCurve)
    qCurve = np.power(curve + delta, 1 - eta)/(1 - eta)
    qCurveAntithetique = np.power(curve + delta, 1 - eta)/(1 - eta)
    
    # f and g vectors 
    fValues = fZero + (1 - fZero)*np.exp(- gamma * np.arange(len(forwardCurve)))
    gValues = ((a + b * np.arange(1, len(forwardCurve) +1)) * np.exp(- c * np.arange(1, len(forwardCurve) +1))) + d
    
    # Product of g and betas (Since they have the same index)
    gBeta1 = gValues * betas[0, :tenor+expiry-1]
    gBeta2 = gValues * betas[1, :tenor+expiry-1]

    gaussians = np.random.standard_normal((timeSteps, 2))
    gaussiansAntithetique = - copy.deepcopy(gaussians)
    
    # These will be important in discouting the forward payoff for each scenario
    deflateur = zeroCouponCurve[0]
    deflateurA = zeroCouponCurve[0]
    DF, DFA = 1, 1
       
    # Loop to construct curves
    for t in range(1, expiry+1):
        
        # Calculate Drift
        mu1 = np.array(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*
                             np.cumsum(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]* np.power(curve + delta, eta)/
                                   (1 + curve)))       
        mu2 = np.array(fValues[t-1]*gBeta2[:len(gBeta2)-t+1] *
                             np.cumsum(fValues[t-1]*gBeta2[:len(gBeta2)-t+1]* np.power(curve + delta, eta)/
                                   (1 + curve)))
        mu = mu1 + mu2
        
        mu1A = np.array(fValues[t-1]*gBeta1[:len(gBeta1)-t+1] *
                             np.cumsum(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]* np.power(curveAntithetique + delta, eta)/
                                   (1 + curveAntithetique)))       
        mu2A = np.array(fValues[t-1]*gBeta2[:len(gBeta2)-t+1] *
                             np.cumsum(fValues[t-1]*gBeta2[:len(gBeta2)-t+1]* np.power(curveAntithetique + delta, eta)/
                                   (1 + curveAntithetique)))
        muA = mu1A + mu2A
        
        fCurveSum = np.power(curve + delta, eta-1)
        fCurveSumA = np.power(curveAntithetique + delta, eta-1)
        
        # Recalculate the qCurve as it should be
        # Construct Curve     
        qCurve = (qCurve
                    + dt * (mu 
                    - 0.5 *eta*  fCurveSum * (np.power(fValues[t-1]*gBeta1[:len(gBeta1)-t+1], 2) + 
                                 np.power(fValues[t-1]*gBeta2[:len(gBeta2)-t+1], 2)))
                    + fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*np.sqrt(dt)*gaussians[t-1, 0] 
                    + fValues[t-1]*gBeta2[:len(gBeta2)-t+1]*np.sqrt(dt)*gaussians[t-1, 1] )
        
        qCurveAntithetique = (qCurveAntithetique
            + dt * (muA 
            - 0.5 *eta* fCurveSumA * (np.power(fValues[t-1]*gBeta1[:len(gBeta1)-t+1], 2) + 
                         np.power(fValues[t-1]*gBeta2[:len(gBeta2)-t+1], 2)))
            + fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*np.sqrt(dt)*gaussiansAntithetique[t-1, 0] 
            + fValues[t-1]*gBeta2[:len(gBeta2)-t+1]*np.sqrt(dt)*gaussiansAntithetique[t-1, 1] )
        
        # Find the discount factor at each timestep
        deflateur = deflateur * DF
        DF = 1/(1 + (np.power((1-eta)*qCurve[0], 1/(1-eta))) - delta)
        
        qCurve = qCurve[1:] # Delete 1 year from curve
        curve = np.power((1-eta)*qCurve, 1/(1-eta)) - delta
        
        # Find the discount factor at each timestep
        deflateurA = deflateurA * DFA
        DFA = 1/(1 + (np.power((1-eta)*qCurveAntithetique[0], 1/(1-eta))) - delta)
        
        qCurveAntithetique = qCurveAntithetique[1:] # Delete 1 year from curve
        curveAntithetique = np.power((1-eta)*qCurveAntithetique, 1/(1-eta)) - delta
               
        
    # Move from Forward Curve to Swap Forward Rate 
    ZCCurve = np.append([DF], DF * np.cumprod(1/(1+curve)))
    PVBP = np.sum(ZCCurve)
    forwardSwap = (1 - ZCCurve[len(ZCCurve)-1])/PVBP
    PVBPDeflated  = PVBP * deflateur
    payoff = np.maximum(forwardSwap - strike, 0) * PVBPDeflated
    
    
    ZCCurveAntithetique = np.append([DFA], DFA * np.cumprod(1/(1+curveAntithetique)))
    PVBPAntithetique = np.sum(ZCCurveAntithetique)
    forwardSwapAntithetique = (1 - ZCCurveAntithetique[len(ZCCurveAntithetique)-1])/PVBPAntithetique
    PVBPDeflatedA  = PVBPAntithetique * deflateurA
    payoffAntithetique = np.maximum(forwardSwapAntithetique - strike, 0) * PVBPDeflatedA
    
    payoff = (payoff + payoffAntithetique)/2
    
    return(payoff)

# Classic Monte Carlo with Moodys Gaussians
###################################################
def BHUnitMonteCarloPayoff(forwardCurve, expiry, tenor, fZero, gamma, a, b, c, d, eta, betas, delta, dZ1, dZ2, strike):
    '''Simulates a single entire swaption scenario using Moodys gaussians.
    Inputs
    ========
    forwardCurve = Curve intended to be projected. Select the curve length depending on the maximum maturity intended for projection.
    
    fZero, gamma, a, b, c, d = parameters for the volatility parametrization
    
    dZ1, dZ2 = Gaussians obtained from the Moodys tool
    '''
    forwardCurve = np.array(forwardCurve)[:tenor+expiry-1]
    timeSteps = len(forwardCurve) 
    dt = 1
      
    # Set the forward Curve
    curve = np.array(forwardCurve)
    qCurve = np.power(curve + delta, 1 - eta)/(1 - eta)
    
    # f and g vectors 
    fValues = fZero + (1 - fZero)*np.exp(- gamma * np.arange(len(forwardCurve)))
    gValues = ((a + b * np.arange(1, len(forwardCurve) +1)) * np.exp(- c * np.arange(1, len(forwardCurve) +1))) + d
    
    # Product of g and betas (Since they have the same index)
    gBeta1 = gValues * betas[0, :tenor+expiry-1]
    gBeta2 = gValues * betas[1, :tenor+expiry-1]

    gaussians = np.array([dZ1[:tenor+expiry-1], dZ2[:tenor+expiry-1]]).transpose()
    
    deflateur = zeroCouponCurve[0]
    DF = 1
    
    # Loop to construct curves
    for t in range(1, expiry+1):
        
        # Calculate Drift
        mu1 = np.array(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*
                             np.cumsum(fValues[t-1]*gBeta1[:len(gBeta1)-t+1]* np.power(curve + delta, eta)/
                                   (1 + curve)))       
        mu2 = np.array(fValues[t-1]*gBeta2[:len(gBeta2)-t+1] *
                             np.cumsum(fValues[t-1]*gBeta2[:len(gBeta2)-t+1]* np.power(curve + delta, eta)/
                                   (1 + curve)))
        mu = mu1 + mu2
        
        fCurveSum = np.power(curve + delta, eta-1)
        
        # Recalculate the qCurve as it should be
        # Construct Curve     
        qCurve = (qCurve
                    + dt * (mu 
                    - 0.5 *eta*  fCurveSum * (np.power(fValues[t-1]*gBeta1[:len(gBeta1)-t+1], 2) + 
                                 np.power(fValues[t-1]*gBeta2[:len(gBeta2)-t+1], 2)))
                    + fValues[t-1]*gBeta1[:len(gBeta1)-t+1]*np.sqrt(dt)*gaussians[t-1, 0] 
                    + fValues[t-1]*gBeta2[:len(gBeta2)-t+1]*np.sqrt(dt)*gaussians[t-1, 1] )
              
        # Find the discount factor at each timestep
        deflateur = deflateur * DF
        DF = 1/(1 + (np.power((1-eta)*qCurve[0], 1/(1-eta))) - delta)
        
        qCurve = qCurve[1:] # Delete 1 year from curve
        curve = np.power((1-eta)*qCurve, 1/(1-eta)) - delta
        
    # Move from Forward Curve to Swap Forward Rate 
    meanZCCurve = np.append([DF], DF * np.cumprod(1/(1+curve)))
    meanPVBP = np.sum(meanZCCurve)
    forwardSwap = (1 - meanZCCurve[len(meanZCCurve)-1])/meanPVBP
      
    PVBPDeflated  = meanPVBP * deflateur
    payoff = np.maximum(forwardSwap - strike, 0) * PVBPDeflated
    
    return(payoff)


'''
6) MONTE CARLO PRICER
==================================================================================================='''
def MonteCarloPricer(forwardCurve, expiry, tenor, fZero, gamma, a, b, c, d, eta, betas, delta, simulations):
    '''Loop to simulate swaption price scenarios using Python gaussians.
        Based on the UnitMonteCarloPayoff function
    
    Inputs
    ========
    forwardCurve = Curve intended to be projected. Select the curve length depending on the maximum maturity intended for projection.

    fZero, gamma, a, b, c, d = parameters for the volatility parametrization'''
    start = time()
    payOff= []
    strike = forwardSwapRateCalib(expiry, tenor)
    
    for simu in np.arange(int(simulations/2)):
        np.random.seed(simu)
        payOff.append(UnitMonteCarloPayoff(forwardCurve, expiry, tenor, fZero, gamma, a, b, c, d, eta, 
                                      betas, delta, strike))

    
    price = np.nanmean(payOff) 
    end = time()
    
    timetaken = round((end -start)/60, 2)
    print(str(expiry)+' X '+str(tenor)+ ' Simulation Completed: '+str(timetaken) + ' minutes')
    
    return(price)


# Moody's Monte Carlo Pricer
def MonteCarloPricerBH(forwardCurve, expiry, tenor, fZero, gamma, a, b, c, d, eta, betas, delta, gaussians1, gaussians2):
    '''Loop to simulate swaption price scenarios using Moodys gaussians.
        Based on the BHUnitMonteCarloPayoff function
    
    Inputs
    ========
    forwardCurve = Curve intended to be projected. Select the curve length depending on the maximum maturity intended for projection.
  
    fZero, gamma, a, b, c, d = parameters for the volatility parametrization'''
    start = time()
    payOff= []
    strike = forwardSwapRateCalib(expiry, tenor)
    
    for simu in np.arange(int(len(gaussians1))):
        payOff.append(BHUnitMonteCarloPayoff(forwardCurve, expiry, tenor, fZero, gamma, a, b, c, d, eta, 
                                      betas, delta, gaussians1[simu], gaussians2[simu], strike))

    end = time()    
    price = np.nanmean(payOff)
    
    timetaken = round((end -start)/60, 2)
    print(str(expiry)+' X '+str(tenor)+ ' Simulation Completed: '+str(timetaken) + ' minutes')
    
    return(price)


'''
**********************************************************************************************************
*                                   IV) MARTINGALE TEST                                                  *
**********************************************************************************************************
'''

'''1) MOODYS MARTINGALE TEST
**********************************'''
def MGTest(simulatedCurves, deflateurs):
    '''MARTINGALE TESTS
    ==========================
    The objective is to run all the Zero Coupon martingale tests in one function. The only inputs are all the scenarios.
    
    Martingale Tests
    =================
    1) Zero Coupon Test
        - Forward Scenarios arranged in Time to Maturity format (TTM)
        - Extraction of the first simulation
        - Zero Coupons transformed from the forwards and arranged in TTM
        
    2) Discounted (Deflated) Zero Coupons
        
    3) Comparison of average discounting factor with the corresponding ZC Curve at time 0
        
    NOTE
    ======
    All results are saved in the Results.xlsx file in the Outputs folder.
    '''
        
    '''FORWARDS TO ZERO COUPONS TO DISCOUNT FACTORS
    ================================================'''
    # Obtain Discount Factors for each scenario
    DF = [np.append(zeroCouponCurve[0],
                    zeroCouponCurve[0]*np.cumprod(df)) for df in deflateurs]
    
    # Average Deflateur
    AvgDeflateur = np.mean(DF, axis = 0) 
    
    # Obtain the Zero Coupon Bonds
    zCoupons = copy.deepcopy(simulatedCurves)
    zCouponsTilde = copy.deepcopy(simulatedCurves)
    
    # Calculate Zero Coupons and Deflated Zero Coupons
    ###################################################################################"
    for i in range(len(simulatedCurves)):
        for j in range(len(simulatedCurves[0])):
            zCoupons[i][j] = np.cumprod(1/(1+simulatedCurves[i][j]))
            
            # Multiply each ZC by the discount factor to obtain Discount factors at each timestep
            zCouponsTilde[i][j] = zCoupons[i][j]*DF[i][j]
    
    # Calculate Average Deflated ZC
    ###################################################################################
    # Copy the triangle shape
    resMean = copy.deepcopy(zCoupons[0])
    
    # For calculation of the mean Deflated ZCoupons
    for i in range(len(resMean)):
        resMean[i] = [0]*len(resMean[i])
    
    # Calculate the mean Deflated ZCoupons    
    for i in range(len(zCouponsTilde)):
        for j in range(len(zCouponsTilde[0])):
            resMean[j] = np.nansum([resMean[j], zCouponsTilde[i][j]], axis = 0)
            
    # Final Mean len(simulatedCurves) = number of simulations        
    zCouponTildeAvg = [i/len(simulatedCurves) for i in resMean]

    # Average of ZC Simulations
    ###################################################################################"
    # Copy the shape of the 1st triangle of the ZCoupons simulations
    zcMean = copy.deepcopy(zCoupons[0])
    
    # For calculation of the mean ZCoupons
    for i in range(len(resMean)):
        zcMean[i] = [0]*len(zcMean[i])
    
    # Calculate the mean ZCoupon len(simulatedCurves) = number of simulations
    for i in range(len(zCouponsTilde)):
        for j in range(len(zCouponsTilde[0])):
            zcMean[j] = np.nansum([zcMean[j], zCoupons[i][j]], axis = 0)
    
    zCouponAvg = [i/len(simulatedCurves) for i in zcMean]
    
    '''CONVERT TRIANGLE TO RECTANGLE
    ===================================='''
    # Average deflated ZC (Take the First 30 maturities  for each year projected)
    rectangleZCTildeAvg = [zCouponTildeAvg[i][:30] for i in range(50)]
    
    # First Simulation
    ZCSimu1 = [zCoupons[0][i][:30] for i in range(len(zCoupons[0]))]
    forwardsSimu1 = [simulatedCurves[0][i][:30] for i in range(len(simulatedCurves[0]))]

    # Error Calculation
    errorsMGTest = [abs((rectangleZCTildeAvg[i]/zeroCouponCurve[1+i:31+i]) -1)
                                for i in range(len(rectangleZCTildeAvg))]

    '''WRITE RESULTS TO EXCEL
    ===================================='''
    try:
        #Open the file
        book = load_workbook('Outputs\\Results.xlsx')
        writer = pd.ExcelWriter('Outputs\\Results.xlsx', engine = 'openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        
        
        'Martingale Test'
        for j in range(len(rectangleZCTildeAvg)):
            # 1st simulation of Forward to ZC transformation
            pd.DataFrame(ZCSimu1[j]).transpose().to_excel(writer, "Fwd to ZC-Scenario1",
                                                  startcol = 4, # Excel Column E
                                                  startrow = 3+j, # Line 3
                                                index = False, 
                                                 header = False)
            # 1st Simulation of Forwards
            pd.DataFrame(forwardsSimu1[j]).transpose().to_excel(writer, "Fwd to ZC-Scenario1",
                                      startcol = 4, # Excel Column E
                                      startrow = 58+j, # Line 59
                                    index = False, 
                                     header = False)
                
            # Save the results of the ZC Martingale Test
            pd.DataFrame(rectangleZCTildeAvg[j]).transpose().to_excel(writer, "Martingale Test",
                                                      startcol = 8, # Excel Column I
                                                      startrow = 58+j, # Line 59
                                                    index = False, 
                                                     header = False)
            
            # Save the errors between the ZC Martingale Test and avg deflateur
            pd.DataFrame(errorsMGTest[j]).transpose().to_excel(writer, "Martingale Test",
                                                      startcol = 8, # Excel Column I
                                                      startrow = 3+j, # Excel Line 4
                                                    index = False, 
                                                     header = False)  
        
        pd.DataFrame(AvgDeflateur[:50]).to_excel(writer, "Martingale Test",
                                                  startcol = 5, # Column F
                                                  startrow = 3, # Line 59
                                                index = False, 
                                                 header = False)
    
        pd.DataFrame(zeroCouponCurve[:80]).to_excel(writer, "Martingale Test",
                                                  startcol = 4, # Column E
                                                  startrow = 3, # Line 59
                                                index = False, 
                                                 header = False)
        writer.save()
       
        print("")
        print('ZC MARTINGALE TEST COMPLETED.')
        print('-------------------------------')
        print('Please find results in: ' + str(os.getcwd()) + '\\Outputs\\Results.xlsx' )
        print("")
 
    except:
        print("")
        print("ERROR ENCOUNTERED")
        print("-------------------")
        print('- Please ensure the following file is closed.')
        print('- If it is properly closed, then please delete the file and create a new blank file.')
        print('- To confirm, please check the following file:')
        print('           '+str(os.getcwd()) + '\Output\Results.xlsx')
  


'''MARKET CONSISTENCY TEST (MOODYS)
============================================'''
def fullMarketConsistencyTestMoodys(forwardCurve, fZero, gamma, a, b, c, d, eta, betas, delta,
                                    browniensBH1,browniensBH2):
    '''MARKET CONSISTENCY TESTS
    ==================================
    The objective is to run all the swaption martingale tests in one function. The inputs are the calibration outputs i.e.
    
    Inputs
    ===========
    1) Calibration parameters (fZero, gamma, a, b, c, d), 
    2) Metaparameters(eta, delta)
    3) forwardCurve
    4) browniens: gaussians used in Monte Carlo Pricing
    
    Swaption tests
    =================
    1) SigmaAlphaBeta Function
        - Exports all the sigmaAlphaBeta functions for each tenor and maturity used
        - Should be the same for both Chi Square and Hagan Pricing
        
    2) Chi Square Swaption Pricing
        - Pricing Swaptions using the closed form Chi Square Function (Andreasen and Andersen (2002))
        - Done both using the Python and R (Note the potential pitfalls of the Python implementation)
        
    3) Monte Carlo Pricing
        - Gaussians used are obtained from Moodys scenarios.
        - Volatility and Prices obtained
    
    4) Bachelier Pricing and Errors.
    
    NOTE
    ======
    All results are saved in the Results.xlsx file in the Outputs folder.
    '''
    try:
        # 1) SigmaAlphaBeta Function
        ################################################################
        sigmaAlphaBetaDataVect = sigmaAlphaBetaPCAVect(expiriesBachelier, tenorsBachelier, 
                                                       fZero, gamma, a, b, c, d, eta, delta)
        
        # Construct Data
        sigmaAlphaBetaData = pd.DataFrame([expiriesBachelier, tenorsBachelier, sigmaAlphaBetaDataVect]).transpose()
        sigmaAlphaBetaData.columns =  ['Expiry', 'Tenor', 'Error']
        sigmaAlphaBetaData = sigmaAlphaBetaData.pivot(index = 'Expiry', columns = 'Tenor', values = 'Error')
    
    except:
        print("")
        print("PRICING ERROR ENCOUNTERED:")
        print("--------------------------")
        print("Possible error encountered in the SigmaAlphaBeta Function.")
 
        
        # 2) Pricing Function R
        ################################################################
    try:
        prixChiSquaredAlphaBetaDataVect = calibrationChiSquarePayerVect(1, strikesLMMPlus, strikesLMMPlus, 
                                                                    expiriesBachelier, tenorsBachelier,
                                                                    fZero, gamma, a, b, c, d, eta, delta)
        
        volChiSquareAlphaBetaVect = volNormalATMFunctionVect(expiriesBachelier,tenorsBachelier, 
                                                             prixChiSquaredAlphaBetaDataVect)
        
        # Construct Data
        prixChiSquaredAlphaBetaData = pd.DataFrame([expiriesBachelier, tenorsBachelier, 
                                                    prixChiSquaredAlphaBetaDataVect]).transpose()
        prixChiSquaredAlphaBetaData.columns =  ['Expiry', 'Tenor', 'Weights']
        prixChiSquaredAlphaBetaData = prixChiSquaredAlphaBetaData.pivot(index = 'Expiry', 
                                                                        columns = 'Tenor', values = 'Weights')
        
        volChiSquareAlphaBetaData = pd.DataFrame([expiriesBachelier, 
                                        tenorsBachelier, volChiSquareAlphaBetaVect]).transpose()
        volChiSquareAlphaBetaData.columns =  ['Expiry', 'Tenor', 'Weights']
        volChiSquareAlphaBetaData = volChiSquareAlphaBetaData.pivot(index = 'Expiry', 
                                                                columns = 'Tenor', values = 'Weights')
        
    except:
        print("")
        print("PRICING ERROR ENCOUNTERED:")
        print("--------------------------")
        print("Possible error encountered in the ChiSquarePayer or volNormalATM Functions.")
        
        # 3) Pricing Function Hagan Function 
        ################################################################
    try:
        blackEquivalent = sigma_CEV_Hagan_Vect(expiriesBachelier, tenorsBachelier, 
                                               fZero, gamma, a, b, c, d, eta, delta)
        prixHaganAlphaBetaDataVect = blackVectCalib(expiriesBachelier, tenorsBachelier, blackEquivalent, delta)
        volHaganAlphaBetaVect = volNormalATMFunctionVect(expiriesBachelier,
                                                        tenorsBachelier, prixHaganAlphaBetaDataVect)
        
        # Construct Data
        prixHaganAlphaBetaData = pd.DataFrame([expiriesBachelier, tenorsBachelier, 
                                               prixHaganAlphaBetaDataVect]).transpose()
        prixHaganAlphaBetaData.columns =  ['Expiry', 'Tenor', 'Weights']
        prixHaganAlphaBetaData = prixHaganAlphaBetaData.pivot(index = 'Expiry', columns = 'Tenor', values = 'Weights')
        
        volHaganAlphaBetaData = pd.DataFrame([expiriesBachelier, tenorsBachelier, volHaganAlphaBetaVect]).transpose()
        volHaganAlphaBetaData.columns =  ['Expiry', 'Tenor', 'Weights']
        volHaganAlphaBetaData = volHaganAlphaBetaData.pivot(index = 'Expiry', columns = 'Tenor', values = 'Weights')
    
    except:
        print("")
        print("PRICING ERROR ENCOUNTERED:")
        print("--------------------------")
        print("Possible error encountered in the sigmaCEVHagan or volNormalATM Functions.")      
    
    
    # 4) Pricing Monte Carlo
    ################################################################
    # Import data from BH Brownians
    try:        
        tauxMCABData = [MonteCarloPricerBH(forwardCurve, expiry, tenor, fZero, gamma, a, b, c, d, 
                                 eta, betas, delta, browniensBH1, browniensBH2) for 
                 expiry, tenor in list(zip(expiriesBachelier, tenorsBachelier))]
        
        tauxMCData = np.array(tauxMCABData).flatten()
        
        # Construct Data
        tauxMCData = pd.DataFrame([expiriesBachelier, tenorsBachelier, tauxMCData]).transpose()
        tauxMCData.columns =  ['Expiry', 'Tenor', 'Weights']
        tauxMCData = tauxMCData.pivot(index = 'Expiry', columns = 'Tenor', values = 'Weights')
        
    except:
        print("")
        print("MONTE CARLO PRICING ERROR ENCOUNTERED:")
        print("-----------------------------------------")        
        print("Please ensure that the Moody's Brownian input is in the correct format.")
        print('- To confirm, please check the following file:')
        print('           '+str(os.getcwd()) + '\BH Browniens.xlsx')
    
    # 5) Volatilités Monte Carlo
    #################################################################
    volMCDataVector = volNormalVect(expiriesBachelier, tenorsBachelier, tauxMCABData)
    
    volMCData = pd.DataFrame([expiriesBachelier, tenorsBachelier, volMCDataVector]).transpose()
    volMCData.columns =  ['Expiry', 'Tenor', 'Weights']
    volMCData = volMCData.pivot(index = 'Expiry', columns = 'Tenor', values = 'Weights')
    
    
    
    # 6) Prepare data for Bachelier Pricing
    ###############################################################
    
    # Normal Prices Data
    ###############################################################
    normalPricesTestData = pd.DataFrame([expiriesBachelier, tenorsBachelier, normalPricesLMMPlus]).transpose()
    normalPricesTestData.columns =  ['Expiry', 'Tenor', 'Weights']
    normalPricesTestData = normalPricesTestData.pivot(index = 'Expiry', columns = 'Tenor', values = 'Weights')
    
    # Normal Volatility Data
    ###############################################################
    volatilitiesLMMPlus.columns =  ['Expiry', 'Tenor', 'Weights']
    normalVolTestData = volatilitiesLMMPlus.pivot(index = 'Tenor', columns = 'Expiry', values = 'Weights') 
    
    
    # 7) Calibration Errors
    ###############################################################
    calibrationErrors =np.abs((normalPricesLMMPlus - prixChiSquaredAlphaBetaDataVect)/normalPricesLMMPlus)
    calibrationErrorsData = pd.DataFrame([expiriesBachelier, tenorsBachelier, calibrationErrors]).transpose()
    calibrationErrorsData.columns =  ['Expiry', 'Tenor', 'Value']
    calibrationErrorsData = calibrationErrorsData.pivot(index = 'Expiry', columns = 'Tenor', values = 'Value')
    
    try:
        '''SAVE RESULTS IN EXCEL
        ==========================================================================='''
        book = load_workbook('Outputs\\Results.xlsx')
        writer = pd.ExcelWriter('Outputs\\Results.xlsx', engine = 'openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        
        normalPricesTestData.to_excel(writer, "Calibration",
                                          startcol = 5, 
                                          startrow = 20, 
                                        index = True, 
                                         header = True)
        
        normalVolTestData.to_excel(writer, "Calibration",
                                          startcol = 5, 
                                          startrow = 3, 
                                        index = True, 
                                         header = True)
        
        calibrationErrorsData.to_excel(writer, "Calibration",
                                          startcol = 5, 
                                          startrow = 37, 
                                        index = True, 
                                         header = True)
        
        sigmaAlphaBetaData.to_excel(writer, "Pricing",
                                          startcol = 5, 
                                          startrow = 2, 
                                        index = True, 
                                         header = True)
        
        prixHaganAlphaBetaData.to_excel(writer, "Pricing",
                                          startcol = 5, 
                                          startrow = 17, 
                                        index = True, 
                                         header = True)
        
        prixChiSquaredAlphaBetaData.to_excel(writer, "Pricing",
                                          startcol = 21, 
                                          startrow = 17, 
                                        index = True, 
                                          header = True)
        
        volHaganAlphaBetaData.to_excel(writer, "Pricing",
                                          startcol = 5, 
                                          startrow = 32, 
                                        index = True, 
                                          header = True)
        
        volChiSquareAlphaBetaData.to_excel(writer, "Pricing",
                                          startcol = 21, 
                                          startrow = 32, 
                                        index = True, 
                                          header = True)
        
        normalVolTestData.to_excel(writer, "Monte Carlo",
                                          startcol = 1, 
                                          startrow = 2, 
                                          index = True, 
                                          header = True)
        
        tauxMCData.to_excel(writer, "Monte Carlo",
                                          startcol = 1, 
                                          startrow = 17, 
                                          index = True, 
                                          header = True)
        
        volMCData.to_excel(writer, "Monte Carlo",
                                          startcol = 1, 
                                          startrow = 32, 
                                          index = True, 
                                          header = True)
        
        # Save all the parameters
        pd.DataFrame([fZero, gamma, a, b, c, d, eta,delta]).to_excel(writer, "Pricing",
                                          startcol = 1, 
                                          startrow = 2, 
                                          index = False, 
                                          header = False)
        
        # Timestamps
        pd.DataFrame([strftime("%H:%M %d/%m/%Y", localtime())]).to_excel(writer, "Pricing",
                                          startcol = 0, 
                                          startrow = 13, 
                                          index = False, 
                                          header = False)
        
        # Save all the parameters
        pd.DataFrame([fZero, gamma, a, b, c, d, eta,delta]).to_excel(writer, "Calibration",
                                          startcol = 1, 
                                          startrow = 22, 
                                          index = False, 
                                          header = False)
        
        # Timestamps
        pd.DataFrame([strftime("%H:%M %d/%m/%Y", localtime())]).to_excel(writer, "Calibration",
                                          startcol = 0, 
                                          startrow = 33, 
                                          index = False, 
                                          header = False)
        
        writer.save()
    except:
        print("")
        print("ERROR ENCOUNTERED")
        print("-------------------")
        print('- Please ensure the following file is closed.')
        print('- If it is properly closed, then please delete the file and create a new blank file.')
        print('- To confirm, please check the following file:')
        print('           '+str(os.getcwd()) + '\Outputs\Results.xlsx')


'''2) PYTHON MARTINGALE TEST
**********************************'''  

'''MARKET CONSISTENCY TEST (MOODYS)
============================================'''
def fullMarketConsistencyTest(forwardCurve, fZero, gamma, a, b, c, d, eta, betas, delta, simulations):
    '''MARKET CONSISTENCY TESTS
    ==================================
    The objective is to run all the swaption martingale tests in one function. The inputs are the calibration outputs i.e.
    
    Inputs
    ===========
    1) Calibration parameters (fZero, gamma, a, b, c, d), 
    2) Metaparameters(eta, delta)
    3) forwardCurve
    
    Swaption tests
    =================
    1) SigmaAlphaBeta Function
        - Exports all the sigmaAlphaBeta functions for each tenor and maturity used
        - Should be the same for both Chi Square and Hagan Pricing
        
    2) Chi Square Swaption Pricing
        - Pricing Swaptions using the closed form Chi Square Function (Andreasen and Andersen (2002))
        - Done both using the Python and R (Note the potential pitfalls of the Python implementation)
        
    3) Monte Carlo Pricing
        - Gaussians used are obtained from Python scenarios.
        - Volatility and Prices obtained
    
    4) Bachelier Pricing and Errors.
    
    NOTE
    ======
    All results are saved in the Results.xlsx file in the Outputs folder.
    '''    
    try:
        # 1) SigmaAlphaBeta Function
        ################################################################
        sigmaAlphaBetaDataVect = sigmaAlphaBetaPCAVect(expiriesBachelier, tenorsBachelier, 
                                                       fZero, gamma, a, b, c, d, eta, delta)
        
        # Construct Data
        sigmaAlphaBetaData = pd.DataFrame([expiriesBachelier, tenorsBachelier, sigmaAlphaBetaDataVect]).transpose()
        sigmaAlphaBetaData.columns =  ['Expiry', 'Tenor', 'Error']
        sigmaAlphaBetaData = sigmaAlphaBetaData.pivot(index = 'Expiry', columns = 'Tenor', values = 'Error')
    
    except:
        print("")
        print("PRICING ERROR ENCOUNTERED:")
        print("--------------------------")
        print("Possible error encountered in the SigmaAlphaBeta Function.")
        sys.exit(1)
        
        # 2) Pricing Function R
        ################################################################
    try:
        prixChiSquaredAlphaBetaDataVect = calibrationChiSquarePayerVect(1, strikesLMMPlus, strikesLMMPlus, 
                                                                    expiriesBachelier, tenorsBachelier,
                                                                    fZero, gamma, a, b, c, d, eta, delta)
        
        volChiSquareAlphaBetaVect = volNormalATMFunctionVect(expiriesBachelier,tenorsBachelier, 
                                                             prixChiSquaredAlphaBetaDataVect)
        
        # Construct Data
        prixChiSquaredAlphaBetaData = pd.DataFrame([expiriesBachelier, tenorsBachelier, 
                                                    prixChiSquaredAlphaBetaDataVect]).transpose()
        prixChiSquaredAlphaBetaData.columns =  ['Expiry', 'Tenor', 'Weights']
        prixChiSquaredAlphaBetaData = prixChiSquaredAlphaBetaData.pivot(index = 'Expiry', 
                                                                        columns = 'Tenor', values = 'Weights')
        
        volChiSquareAlphaBetaData = pd.DataFrame([expiriesBachelier, 
                                        tenorsBachelier, volChiSquareAlphaBetaVect]).transpose()
        volChiSquareAlphaBetaData.columns =  ['Expiry', 'Tenor', 'Weights']
        volChiSquareAlphaBetaData = volChiSquareAlphaBetaData.pivot(index = 'Expiry', 
                                                                columns = 'Tenor', values = 'Weights')
        
    except:
        print("")
        print("PRICING ERROR ENCOUNTERED:")
        print("--------------------------")
        print("Possible error encountered in the ChiSquarePayer or volNormalATM Functions.")
        sys.exit(1)
        
        # 3) Pricing Function Hagan Function 
        ################################################################
    try:
        blackEquivalent = sigma_CEV_Hagan_Vect(expiriesBachelier, tenorsBachelier, 
                                               fZero, gamma, a, b, c, d, eta, delta)
        prixHaganAlphaBetaDataVect = blackVectCalib(expiriesBachelier, tenorsBachelier, blackEquivalent, delta)
        volHaganAlphaBetaVect = volNormalATMFunctionVect(expiriesBachelier,
                                                        tenorsBachelier, prixHaganAlphaBetaDataVect)
        
        # Construct Data
        prixHaganAlphaBetaData = pd.DataFrame([expiriesBachelier, tenorsBachelier, 
                                               prixHaganAlphaBetaDataVect]).transpose()
        prixHaganAlphaBetaData.columns =  ['Expiry', 'Tenor', 'Weights']
        prixHaganAlphaBetaData = prixHaganAlphaBetaData.pivot(index = 'Expiry', columns = 'Tenor', values = 'Weights')
        
        volHaganAlphaBetaData = pd.DataFrame([expiriesBachelier, tenorsBachelier, volHaganAlphaBetaVect]).transpose()
        volHaganAlphaBetaData.columns =  ['Expiry', 'Tenor', 'Weights']
        volHaganAlphaBetaData = volHaganAlphaBetaData.pivot(index = 'Expiry', columns = 'Tenor', values = 'Weights')
    
    except:
        print("")
        print("PRICING ERROR ENCOUNTERED:")
        print("--------------------------")
        print("Possible error encountered in the sigmaCEVHagan or volNormalATM Functions.") 
        sys.exit(1)
    
    
    # 4) Pricing Monte Carlo
    ################################################################
    # Import data from BH Brownians
    try: 
        tauxMCABData = [MonteCarloPricer(forwardCurve, expiry, tenor, 
                                         fZero, gamma, a, b, c, d, eta, betas, delta, simulations) for 
                 expiry, tenor in list(zip(expiriesBachelier, tenorsBachelier))]
        
        tauxMCData = np.array(tauxMCABData).flatten()
        
        # Construct Data
        tauxMCData = pd.DataFrame([expiriesBachelier, tenorsBachelier, tauxMCData]).transpose()
        tauxMCData.columns =  ['Expiry', 'Tenor', 'Weights']
        tauxMCData = tauxMCData.pivot(index = 'Expiry', columns = 'Tenor', values = 'Weights')
        
    except:
        print("")
        print("MONTE CARLO PRICING ERROR ENCOUNTERED:")
        print("-----------------------------------------")        
        sys.exit(1)
    
    # 5) Volatilités Monte Carlo
    #################################################################
    volMCDataVector = volNormalVect(expiriesBachelier, tenorsBachelier, tauxMCABData)
    
    volMCData = pd.DataFrame([expiriesBachelier, tenorsBachelier, volMCDataVector]).transpose()
    volMCData.columns =  ['Expiry', 'Tenor', 'Weights']
    volMCData = volMCData.pivot(index = 'Expiry', columns = 'Tenor', values = 'Weights')
    
    
    
    # 6) Prepare data for Bachelier Pricing
    ###############################################################
    
    # Normal Prices Data
    ###############################################################
    normalPricesTestData = pd.DataFrame([expiriesBachelier, tenorsBachelier, normalPricesLMMPlus]).transpose()
    normalPricesTestData.columns =  ['Expiry', 'Tenor', 'Weights']
    normalPricesTestData = normalPricesTestData.pivot(index = 'Expiry', columns = 'Tenor', values = 'Weights')
    
    # Normal Volatility Data
    ###############################################################
    volatilitiesLMMPlus.columns =  ['Expiry', 'Tenor', 'Weights']
    normalVolTestData = volatilitiesLMMPlus.pivot(index = 'Tenor', columns = 'Expiry', values = 'Weights') 
    
    
    # 7) Calibration Errors
    ###############################################################
    calibrationErrors =np.abs((normalPricesLMMPlus - prixChiSquaredAlphaBetaDataVect)/normalPricesLMMPlus)
    calibrationErrorsData = pd.DataFrame([expiriesBachelier, tenorsBachelier, calibrationErrors]).transpose()
    calibrationErrorsData.columns =  ['Expiry', 'Tenor', 'Value']
    calibrationErrorsData = calibrationErrorsData.pivot(index = 'Expiry', columns = 'Tenor', values = 'Value')
    
    try:
        '''SAVE RESULTS IN EXCEL
        ==========================================================================='''
        book = load_workbook('Outputs\\Results.xlsx')
        writer = pd.ExcelWriter('Outputs\\Results.xlsx', engine = 'openpyxl')
        writer.book = book
        writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
        
        normalPricesTestData.to_excel(writer, "Calibration",
                                          startcol = 5, 
                                          startrow = 20, 
                                        index = True, 
                                         header = True)
        
        normalVolTestData.to_excel(writer, "Calibration",
                                          startcol = 5, 
                                          startrow = 3, 
                                        index = True, 
                                         header = True)
        
        calibrationErrorsData.to_excel(writer, "Calibration",
                                          startcol = 5, 
                                          startrow = 37, 
                                        index = True, 
                                         header = True)
        
        sigmaAlphaBetaData.to_excel(writer, "Pricing",
                                          startcol = 5, 
                                          startrow = 2, 
                                        index = True, 
                                         header = True)
        
        prixHaganAlphaBetaData.to_excel(writer, "Pricing",
                                          startcol = 5, 
                                          startrow = 17, 
                                        index = True, 
                                         header = True)
        
        prixChiSquaredAlphaBetaData.to_excel(writer, "Pricing",
                                          startcol = 21, 
                                          startrow = 17, 
                                        index = True, 
                                          header = True)
        
        volHaganAlphaBetaData.to_excel(writer, "Pricing",
                                          startcol = 5, 
                                          startrow = 32, 
                                        index = True, 
                                          header = True)
        
        volChiSquareAlphaBetaData.to_excel(writer, "Pricing",
                                          startcol = 21, 
                                          startrow = 32, 
                                        index = True, 
                                          header = True)
        
        normalVolTestData.to_excel(writer, "Monte Carlo",
                                          startcol = 1, 
                                          startrow = 2, 
                                          index = True, 
                                          header = True)
        
        tauxMCData.to_excel(writer, "Monte Carlo",
                                          startcol = 1, 
                                          startrow = 17, 
                                          index = True, 
                                          header = True)
        
        volMCData.to_excel(writer, "Monte Carlo",
                                          startcol = 1, 
                                          startrow = 32, 
                                          index = True, 
                                          header = True)
        # Timestamps
        pd.DataFrame([strftime("%H:%M %d/%m/%Y", localtime())]).to_excel(writer, "Pricing",
                                          startcol = 0, 
                                          startrow = 13, 
                                          index = False, 
                                          header = False)
        
        # Save all the parameters
        pd.DataFrame([fZero, gamma, a, b, c, d, eta,delta]).to_excel(writer, "Calibration",
                                          startcol = 1, 
                                          startrow = 22, 
                                          index = False, 
                                          header = False)
        
        # Timestamps
        pd.DataFrame([strftime("%H:%M %d/%m/%Y", localtime())]).to_excel(writer, "Calibration",
                                          startcol = 0, 
                                          startrow = 33, 
                                          index = False, 
                                          header = False)
        writer.save()
    except:
        print("")
        print("ERROR ENCOUNTERED")
        print("-------------------")
        print('- Please ensure the following file is closed.')
        print('- If it is properly closed, then please delete the file and create a new blank file.')
        print('- To confirm, please check the following file:')
        print('           '+str(os.getcwd()) + '\Outputs\Results.xlsx')
        sys.exit(1)

'''
**********************************************************************************************************
*                                   V) MISCELLANEOUS                                                     *
**********************************************************************************************************
'''
def plotter(eta, delta, result):       
    weightsPlot = np.array([((tenorMaturity[0] >= 2) & (tenorMaturity[0] <= 30 ) & 
                            (tenorMaturity[1] >= 2) & (tenorMaturity[1] <= 30 )) 
                        for tenorMaturity in tenorMaturities])
    
    strikesCalibPlot = np.array(list(compress(strikes, weightsPlot)))
    expiriesCalibPlot = list(compress(expiries, weightsPlot))
    tenorsCalibPlot = list(compress(tenors, weightsPlot))
    normalPricesCalibPlot = list(compress(normalPrices, weightsPlot))
    
    percentageErrors = np.absolute((calibrationChiSquarePayerVect(1, strikesCalibPlot, 
                        strikesCalibPlot, 
                        expiriesCalibPlot, 
                        tenorsCalibPlot, 
                        result.params['fZero'].value, 
                        result.params['gamma'].value, 
                        result.params['a'].value, 
                        result.params['b'].value,
                        result.params['c'].value,
                        result.params['d'].value,
                        eta,
                        delta)/normalPricesCalibPlot)-1)*100 
    
    percentageErrors = percentageErrors.astype(float)
    
    # Construct Data
    heatmapData = pd.DataFrame([expiriesCalibPlot, tenorsCalibPlot, percentageErrors]).transpose()
    heatmapData.columns =  ['Expiry', 'Tenor', 'Error']
    heatmapData = heatmapData.pivot(index = 'Expiry', columns = 'Tenor', values = 'Error')
    
    # Constructing the plot
    fig, ax = plt.subplots(figsize = (40, 36), facecolor='#444444')
    
    # Add Title
    title = "Percentage Calibration Errors: Eta = " + str(eta) + " Delta = "+str(delta)
    
    # Set the title and distance from the plo
    ttl = ax.title
    ttl.set_position([0.5, 1.05])
    ax.invert_yaxis()
    ax.set_title(title, color='black', fontsize = 50)
    
    # Define labels
    labels = np.array(percentageErrors.transpose()).reshape(heatmapData.shape)
    
    # Use seaborn 
    plot = sns.heatmap(heatmapData,
                cmap = "viridis",
                cbar= False,
                annot = True,
               linewidths = 0.3, 
               ax = ax,
               annot_kws={'size':40})
    
    plot.axes.xaxis.set_ticks_position('top')
    plt.xlabel('Tenor', fontsize = 40, color = 'black')
    plt.ylabel('Expiry', fontsize = 40, color = 'black')
    
    ax.tick_params(labelcolor='black',
                  labelsize = 38)
    plt.savefig('Output\\Fit\\Delta = '+ str(round(delta, 3))+ ' Eta = ' + str(round(eta, 2)) + ' Calibration ErrorsPCA.png')



# '''CALIBRATION DATA
# ====================='''
# loc = "01 - Calibration Data.xlsx"

# #Obtain ATM Normal Volatility Data
# vol_data = pd.read_excel(loc,
#                         sheet_name = "ATM Swaptions Normal Vol",
#                         skiprows= [0, 2,3],
#                         index_col = "Dates")

# # Decimal Notation
# vol_data = pd.DataFrame.dropna(vol_data, axis = 'rows')/10000

# # Obtain Swap Curve Data
# swap_data = pd.read_excel(loc,
#                         sheet_name = "Swap Rates Data",
#                         skiprows= [0, 1,3],
#                        index_col = "Dates")

# # Decimal Notation
# swap_data = pd.DataFrame.dropna(swap_data, axis = 'rows')/100
# swap_data = swap_data.loc[vol_data.index]

# '''CURVE SETUP
# ================='''
# # Swap Rate maturities
# swapMaturities = np.array(([i for i in range(1,31)] + # Years
#                           [i*5 for i in range(7, 13)])) # 50 years

# ''' SWAP RATE INTERPOLATION'''
# '''Piecewise Linear Interpolation - Used in Bloomberg'''
# # Range of dates for each maturity
# fullMaturities = np.arange(1, 61)
# fullSwapData = pd.DataFrame([np.interp(fullMaturities, 
#                                        swapMaturities, swap_data.loc[date])
#                for date in swap_data.index],
#                     columns = list(range(1, 61)),
#                     index = swap_data.index)

# '''ZERO COUPON BOOTSTRAPPING (Single Curve)'''
# # Construct Zero Coupon Prices for each maturity
# # We already have the first ZC date
# zeroCoupons = []

# # Loop to calculate the ZC prices from swap curves
# for j in range(len(swap_data.index)):
#     curve = np.array(fullSwapData.loc[swap_data.index[j]])
#     zeroCoupons.append([])
    
#     for i in range(0, 60):
#         zeroCoupons[j].append((1 - curve[i] * np.sum(zeroCoupons[j][:i]))/
#                                            (1 + curve[i]))

# # Save zeroCoupons as a dataframe
# zeroCoupons = pd.DataFrame(zeroCoupons,
#                     columns = list(range(1, 61)),
#                     index = swap_data.index)

# '''FORWARD CURVE CONSTRUCTION'''
# #Range of end dates for forwards ie date K for F(t; K, K+1)
# forwardStartDates = np.arange(2, 61)

# forwardCurves = []

# # Loop to calculate forward prices from ZC Prices
# for j in range(len(swap_data.index)):
#     curve = np.array(zeroCoupons.loc[swap_data.index[j]])
#     forwardCurves.append([(curve[i-1]/curve[i])-1 for i in range(1, len(curve))])

        
# # Save forwards as a dataframe
# forwardCurves = pd.DataFrame(forwardCurves,
#                     columns = list(range(1, 60)),
#                     index = swap_data.index)

# '''SELECT THE CURVES WE WANT TO USE
# =====================================
# Based on the date and curves intended for use. (Either EIOPA or Market Curves)
# '''
# testVolData = vol_data.loc['2019-12-31']
# zeroCouponCurve = zeroCoupons.loc['2019-12-31']
# forwardCurve = forwardCurves.loc['2019-12-31']

'''SELECT SWAPTIONS
=====================================

Note that vol data is contained in the testVolData variable. 

A function has been provided to change the data based on the calibration date required.'''

'''FIND STRIKES FOR EACH SWAPTION'''
# # Collect only annual Data
# annualTenorNames = []
# annualTenors = []

# for string in testVolData.index:
#     if 'M' not in string:
#         annualTenorNames.append(string)
#         annualTenors.append(testVolData.loc[string])

# tenorMaturities =  [(int(re.findall('[0-9]+',string)[0]),int(re.findall('[0-9]+',string)[1]) ) 
#                     for string in annualTenorNames]

# expiries = [int(i[0]) for i in tenorMaturities]

# tenors = [int(i[1]) for i in tenorMaturities]

# #Obtain all strikes
# strikes = forwardSwapRateVect(expiries, tenors) 

# # Calculate value of all swaptions
# normalPrices = normalPayerVect(1, strikes, strikes, expiries, tenors,annualTenors)
